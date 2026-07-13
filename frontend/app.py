"""
UNIDO IEU Evaluation Intelligence Platform
4-tab Streamlit frontend:
  Tab 1 — Search & Browse   (filter reports, view lessons/recommendations, Excel export)
  Tab 2 — Synthesis         (RAG search across selected reports)
  Tab 3 — Visualize         (portfolio charts via Plotly)
  Tab 4 — OECD-DAC          (DAC criteria evidence browser + radar chart)
  Admin  — User management  (admin only)
"""

import io
import os
import re
import json
import pathlib
import base64
import httpx
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import streamlit as st
from pathlib import Path

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────

def get_backend_url() -> str:
    try:
        if "BACKEND_URL" in st.secrets:
            return st.secrets["BACKEND_URL"]
    except Exception:
        pass
    from dotenv import load_dotenv
    load_dotenv(Path(__file__).parent.parent / ".env")
    return os.getenv("BACKEND_URL", "http://localhost:8000")

BACKEND_URL = get_backend_url()

st.set_page_config(
    page_title="UNIDO Evaluation Intelligence Platform",
    page_icon="",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# SDG data — official UN colours + names
# ─────────────────────────────────────────────────────────────────────────────

SDG_COLORS = {
    1: "#E5243B", 2: "#DDA63A", 3: "#4C9F38", 4: "#C5192D", 5: "#FF3A21",
    6: "#26BDE2", 7: "#FCC30B", 8: "#A21942", 9: "#FD6925", 10: "#DD1367",
    11: "#FD9D24", 12: "#BF8B2E", 13: "#3F7E44", 14: "#0A97D9", 15: "#56C02B",
    16: "#00689D", 17: "#19486A",
}

SDG_NAMES = {
    1: "No Poverty", 2: "Zero Hunger", 3: "Good Health", 4: "Quality Education",
    5: "Gender Equality", 6: "Clean Water", 7: "Affordable Energy", 8: "Decent Work",
    9: "Industry & Innovation", 10: "Reduced Inequalities", 11: "Sustainable Cities",
    12: "Responsible Consumption", 13: "Climate Action", 14: "Life Below Water",
    15: "Life on Land", 16: "Peace & Justice", 17: "Partnerships",
}

# ─────────────────────────────────────────────────────────────────────────────
# SDG icon images — load from frontend/assets/sdg_01.png … sdg_17.png as base64
# Falls back to CSS colored tile if image file not found
# ─────────────────────────────────────────────────────────────────────────────

_ASSETS_DIR = Path(__file__).parent / "assets"
_SDG_B64: dict = {}

def _load_sdg_images():
    for n in range(1, 18):
        p = _ASSETS_DIR / f"sdg_{n:02d}.png"
        if p.exists():
            _SDG_B64[n] = base64.b64encode(p.read_bytes()).decode()

_load_sdg_images()


def sdg_badge_html(n: int, size: int = 36) -> str:
    """Real PNG icon if available, CSS colored tile as fallback."""
    title = f"SDG {n}: {SDG_NAMES.get(n, '')}"
    if n in _SDG_B64:
        return (
            f'<img src="data:image/png;base64,{_SDG_B64[n]}" '
            f'title="{title}" '
            f'style="width:{size}px;height:{size}px;border-radius:4px;margin:2px;'
            f'vertical-align:middle;object-fit:cover;" />'
        )
    # CSS fallback
    color = SDG_COLORS.get(n, "#888")
    font_size = max(9, size // 3)
    return (
        f'<span title="{title}" '
        f'style="display:inline-flex;align-items:center;justify-content:center;'
        f'width:{size}px;height:{size}px;background:{color};color:white;font-weight:800;'
        f'border-radius:4px;font-size:{font_size}px;margin:2px;vertical-align:middle;'
        f'font-family:Arial,sans-serif;letter-spacing:-0.5px;">{n}</span>'
    )


def sdg_badges_row(sdg_list: list, size: int = 32) -> str:
    return "".join(sdg_badge_html(n, size) for n in sorted(sdg_list) if 1 <= n <= 17)

# ─────────────────────────────────────────────────────────────────────────────
# DAC criteria
# ─────────────────────────────────────────────────────────────────────────────

DAC_CRITERIA = ["relevance", "effectiveness", "efficiency", "impact", "sustainability"]
DAC_COLORS   = ["#009EDB", "#4CAF50", "#FF9800", "#E91E63", "#9C27B0"]
DAC_LABELS   = ["Relevance", "Effectiveness", "Efficiency", "Impact", "Sustainability"]

THEMATIC_AREAS = [
    "Energy Efficiency", "Clean / Renewable Energy", "Climate Action",
    "Circular Economy / Waste Management", "Chemicals & POPs",
    "Industrial Policy & Competitiveness", "Trade & Standards",
    "Agro-Industry & Food Systems", "Water & Environment",
    "Gender & Inclusion", "Digital Innovation",
]

# ── Pilot report metadata (source of truth — overrides backend/Qdrant) ───────
PILOT_METADATA: dict[str, dict] = {
    "UNIDO-100043": {
        "project_id":          "100043",
        "gef_id":              "4114",
        "title":               "Independent Terminal Evaluation: The Project \"Bamboo Processing for Sri Lanka\"",
        "short_title":         "Bamboo Processing for Sri Lanka",
        "year":                2021,
        "country":             "Sri Lanka",
        "region":              "Asia and the Pacific",
        "thematic_category":   "Agro-Industry & Food Systems",
        "secondary_thematic_area": "Circular Economy / Waste Management",
        "report_type":         "Project Evaluation",
        "donor":               "GEF",
        "budget_usd":          2_355_000,
        "evaluation_rating":   2.0,
        "overall_rating_label":"Unsatisfactory",
        "util_rate":           100.8,
        # ── Duration & delay (Sept 2012 – Mar 2021, 13-month overrun) ────────
        "planned_months":      90,
        "actual_months":       103,
        "delay_months":        13,
        "overrun_pct":         14,
        "has_delay":           True,
        "start_year":          2012,
        "planned_end_year":    2020,
        "actual_end_year":     2021,
        "duration_str":        "Sept 2012 – Mar 2021",
        "delay_causes": (
            "A Project Management Office was not established on time, delaying coordination from inception. "
            "Touchwood Investments PLC — primary private partner pledging USD 1.3M — went bankrupt early. "
            "Of USD 21.3M co-financing pledged at design, only UNIDO's USD 100,000 was received. "
            "A USD 600,000 revolving fund sat idle at Hatton National Bank (2014–18) before being "
            "cancelled by the Steering Committee, disrupting financial planning for two years."
        ),
        # ── Year-by-year GEF disbursement (Annex — Budget Execution table) ───
        "disbursement_by_year": {
            2012: 38_884,
            2013: 261_518,
            2014: 597_169,
            2015: 498_978,
            2016: 254_890,
            2017: 241_738,
            2019: 551_592,  # 2018 excluded: revolving fund reversal (accounting entry, not real spend)
            2020: 110_856,
        },
        # ── Co-financing planned vs actually received (Table 13) ──────────────
        "cofinancing_labels":  ["Mahaweli Auth.", "Ministry Industry", "Forest Dept.", "Touchwood PLC", "Bamboo Resources", "UNIDO"],
        "cofinancing_planned": [12_920_000, 1_500_000, 4_377_000, 1_300_000, 1_100_000, 100_000],
        "cofinancing_actual":  [0, 0, 0, 0, 0, 100_000],
        # ── Timeline milestones (Annex 9) ─────────────────────────────────────
        "timeline_events": [
            {"yr": 2012, "lab": "Project start",              "type": "plan"},
            {"yr": 2013, "lab": "Bamboo Growers Assoc. est.", "type": "actual"},
            {"yr": 2014, "lab": "Revolving fund established", "type": "actual"},
            {"yr": 2016, "lab": "Mid-Term Evaluation",        "type": "actual"},
            {"yr": 2018, "lab": "Revolving fund cancelled",   "type": "actual"},
            {"yr": 2019, "lab": "PMO established",            "type": "actual"},
            {"yr": 2021, "lab": "Project completion",         "type": "actual"},
        ],
        # ── Gender: project was gender-blind (rating: Unsatisfactory) ─────────
        "stakeholders_rich": None,
        "gender_rating": "Unsatisfactory",
        "gender_note": (
            "The project had no gender mainstreaming objectives and was effectively gender-blind. "
            "Training participants were predominantly male; no initiatives encouraged female participation "
            "except in bamboo handcraft and design. Outcomes disproportionately favoured males. "
            "Gender mainstreaming is rated Unsatisfactory."
        ),
        # ── Per-criterion DAC ratings (1–6 scale: HS=6, S=5, MS=4, MU=3, U=2, HU=1) ─
        "dac_ratings": {
            "relevance":      4,   # Moderately Satisfactory — relevant to national priorities but design gaps
            "effectiveness":  2,   # Unsatisfactory — core objectives not achieved; value chain did not materialise
            "efficiency":     2,   # Unsatisfactory — 13-month overrun; USD 21M co-financing received = 0
            "impact":         2,   # Unsatisfactory — minimal demonstrable economic or environmental impact
            "sustainability": 2,   # Unsatisfactory — enterprises financially unviable at project close
        },
        "dac_rating_labels": {
            "relevance":      "Moderately Satisfactory",
            "effectiveness":  "Unsatisfactory",
            "efficiency":     "Unsatisfactory",
            "impact":         "Unsatisfactory",
            "sustainability": "Unsatisfactory",
        },
        "sdgs":                [8, 12, 15],
        "thematic_justification": (
            "The project focused on developing the bamboo value chain as an agro-industrial sector, "
            "supporting smallholder farmers and SMEs in the processing and commercialization of "
            "bamboo-based products as a sustainable alternative to timber. Its core activities centred "
            "on agro-industry capacity building, technology transfer for bamboo processing equipment, "
            "and market development for agricultural value-added products."
        ),
        "sdg_justifications": {
            "8":  "The project aimed to create employment and income for smallholder bamboo farmers and processing enterprises in rural Sri Lanka, directly contributing to SDG 8 (Decent Work and Economic Growth) through agro-industrial value chain development.",
            "12": "By promoting bamboo as a renewable, fast-growing alternative to timber and conventional inputs, the project advanced responsible consumption and production patterns (SDG 12) in the construction, handicraft, and packaging industries.",
            "15": "Bamboo cultivation and plantation management conserves soil, reduces erosion, and provides an alternative to unsustainable timber extraction, contributing to SDG 15 (Life on Land) through sustainable forestry and land-use practices in Sri Lanka.",
        },
        "lessons_learned": [
            "Thorough feasibility and market studies must be completed before project design is finalised — the Sri Lanka bamboo project suffered from overestimated market demand and insufficient baseline data on bamboo resource availability.",
            "Private sector buy-in and co-financing commitments should be secured prior to project start. Over-reliance on small-scale farmers without commercial processing partners limited the project's ability to achieve value chain scale.",
            "Adaptive management mechanisms must be built into project design: when market conditions diverge significantly from projections, project teams need explicit authority and resources to redesign activities mid-course.",
            "Technology transfer activities are only effective when accompanied by sustained market linkage support. Providing processing equipment without ensuring buyers and commercial relationships undermined enterprise viability.",
            "Realistic sustainability planning requires honest assessment of cost recovery potential. Processing enterprises could not achieve financial break-even at projected volumes, indicating project design assumptions were overly optimistic.",
        ],
        "recommendations": [
            "UNIDO and GEF should commission a comprehensive post-project market assessment to determine whether commercial bamboo processing viability has improved sufficiently to justify a follow-on phase with restructured value chain support.",
            "The National Bamboo Committee should establish a dedicated market development function, focused on connecting processing enterprises with domestic and export buyers, as a condition for any future phase of bamboo sector support.",
            "Future agro-industry projects in Sri Lanka should require binding private sector co-financing agreements — not in-kind or contingent contributions — before project approval, to ensure genuine commercial interest and shared financial risk.",
            "UNIDO should integrate systematic adaptive management reviews at 18-month intervals in complex agro-industry projects, with explicit decision gates allowing activity redesign when market assumptions prove incorrect.",
            "The Government of Sri Lanka should establish a bamboo sector monitoring system within the Ministry of Agriculture to track plantation area, processing enterprise performance, and market offtake independently of donor project support.",
        ],
    },
    "UNIDO-100321": {
        # ── Core metadata ────────────────────────────────────────────────────
        "project_id":          "100321",
        "gef_id":              "4602",
        "title":               "Independent Terminal Evaluation: Initiation of the HCFC Phase Out in the Republic of Azerbaijan",
        "short_title":         "HCFC Phase Out – Azerbaijan",
        "year":                2021,
        "country":             "Azerbaijan",
        "region":              "Europe and Central Asia",
        "thematic_category":   "Chemicals & POPs",
        "secondary_thematic_area": "Climate Action",
        "report_type":         "Project Evaluation",
        "donor":               "GEF",
        # ── Financials (GEF grant only, Table 5 of Terminal Evaluation) ──────
        "budget_usd":          2_620_000,
        "evaluation_rating":   5.5,
        "overall_rating_label":"Highly Satisfactory",
        "util_rate":           99.97,
        # ── Duration & delay ─────────────────────────────────────────────────
        "planned_months":      48,
        "actual_months":       70,
        "delay_months":        22,
        "overrun_pct":         46,
        "has_delay":           True,
        "start_year":          2015,
        "planned_end_year":    2019,
        "actual_end_year":     2020,
        "duration_str":        "Feb 2015 – Dec 2020",
        "delay_causes": (
            "The project experienced a 22-month overrun beyond the original Dec 2019 closure date. "
            "GEF funds were 93.9% committed by end of 2017, but delays in enterprise recruitment and "
            "HCFC quota enforcement required continued technical assistance through 2020. "
            "The COVID-19 pandemic further delayed the Terminal Evaluation to 2021."
        ),
        # ── Year-by-year GEF disbursement (Table 5, Terminal Evaluation) ─────
        "disbursement_by_year": {
            2015: 180_895,
            2016: 721_822,
            2017: 1_556_805,
            2018: 20_404,
            2019: 33_517,
            2020: 100_897,
            2021: 19_068,
        },
        # ── Co-financing by enterprise (Table 6, Terminal Evaluation) ────────
        "cofinancing_labels":  ["Fayton", "Titan", "TG Chem", "A&K", "Frigo Mkt", "Baku Chinar", "MENR/UNIDO"],
        "cofinancing_planned": [95_000, 95_000, 43_000, 43_000, 43_000, 43_000, 200_000],
        "cofinancing_actual":  [110_000, 88_000, 50_000, 39_000, 48_000, 40_000, 195_000],
        # ── Implementation timeline milestones ───────────────────────────────
        "timeline_events": [
            {"yr": 2015, "lab": "Project start",       "type": "plan"},
            {"yr": 2016, "lab": "Analysers to customs","type": "actual"},
            {"yr": 2017, "lab": "Mid-Term Review",     "type": "actual"},
            {"yr": 2019, "lab": "Zero HCFC reported",  "type": "actual"},
            {"yr": 2020, "lab": "Actual closure",      "type": "actual"},
            {"yr": 2021, "lab": "Terminal Evaluation", "type": "actual"},
        ],
        # ── Stakeholders & gender (Table 14, Terminal Evaluation) ────────────
        "stakeholders_rich": [
            {"name": "CCOC / MENR",      "role": "National implementing partner", "pct_women": 83},
            {"name": "Fayton Ltd.",       "role": "Enterprise: refrigeration",     "pct_women": 25},
            {"name": "Titan Service Ltd.","role": "Enterprise: refrigeration",     "pct_women": 40},
            {"name": "TG Chemical",       "role": "Enterprise: foam sector",       "pct_women": 100},
            {"name": "A&K",               "role": "Enterprise: air conditioning",  "pct_women": 18},
        ],
        "gender_rating": "Partially Mainstreamed",
        "gender_note": (
            "The project had no formal gender equality strategy or dedicated gender mainstreaming budget. "
            "Female employment was tracked at enterprise level — ranging from 18 % (A&K) to 100 % (TG Chemical) "
            "— but was not systematically integrated into project design or M&E frameworks. "
            "Gender is rated Partially Mainstreamed in line with GEF requirements."
        ),
        # ── Per-criterion DAC ratings (1–6 scale: HS=6, S=5, MS=4, MU=3, U=2, HU=1) ─
        "dac_ratings": {
            "relevance":      6,   # Highly Satisfactory — directly aligned to Montreal Protocol obligations
            "effectiveness":  5,   # Satisfactory — HCFC phase-out achieved; capacity built
            "efficiency":     4,   # Moderately Satisfactory — 22-month overrun but GEF funds 99.97% utilised
            "impact":         5,   # Satisfactory — measurable GHG reductions; institutional change sustained
            "sustainability": 4,   # Moderately Likely — licensing system in place; some institutional fragility
        },
        "dac_rating_labels": {
            "relevance":      "Highly Satisfactory",
            "effectiveness":  "Satisfactory",
            "efficiency":     "Moderately Satisfactory",
            "impact":         "Satisfactory",
            "sustainability": "Moderately Likely",
        },
        # ── SDGs ─────────────────────────────────────────────────────────────
        "sdgs":                [12, 13, 17],
        "thematic_justification": (
            "The project directly targeted the phase-out of hydrochlorofluorocarbons (HCFCs) — "
            "ozone-depleting and high global warming potential substances — under the Montreal Protocol "
            "and the Multilateral Fund. This positions it firmly in the Chemicals & POPs thematic area, "
            "involving import/export controls, customs training, refrigerant conversion in enterprises, "
            "and institutional strengthening for chemical substance management in Azerbaijan."
        ),
        "sdg_justifications": {
            "12": "The project phased out HCFCs in the refrigeration, air conditioning, and foam sectors of Azerbaijan, replacing them with ozone- and climate-friendly alternatives, directly advancing SDG 12 (Responsible Consumption and Production) by eliminating harmful substances from industrial processes.",
            "13": "HCFCs are potent greenhouse gases in addition to being ozone-depleting substances. By phasing them out and transitioning enterprises to low-GWP refrigerants, the project contributes to SDG 13 (Climate Action) through measurable reductions in greenhouse gas emissions.",
            "17": "The project was implemented under the Multilateral Fund for the Implementation of the Montreal Protocol — a flagship example of global partnership. It relied on international technology transfer and multi-stakeholder coordination, directly embodying SDG 17 (Partnerships for Goals).",
        },
        "lessons_learned": [
            "Effective HCFC phase-out requires early and sustained investment in customs officer training and enforcement capacity — without this, import controls remain ineffective regardless of regulatory frameworks in place.",
            "Strong government ownership and institutional embedding of phase-out management structures are the most reliable predictors of long-term programme sustainability beyond project closure.",
            "Enterprise recruitment and conversion timelines are consistently underestimated in phase-out projects; building in flexible implementation schedules and contingency for enterprise mobilisation delays significantly improves delivery.",
            "South-South cooperation and regional experience sharing among Montreal Protocol parties accelerates learning and reduces the cost of technology conversion by enabling beneficiary countries to learn from peers who have already undertaken similar transitions.",
            "Verification and monitoring systems for phase-out achievement must be designed to be maintained by national counterparts after project completion; systems that depend on international consultants are not sustainable.",
        ],
        "recommendations": [
            "The Azerbaijan government should sustain the customs enforcement and HCFC monitoring system established under the project by integrating it into routine operations of the State Customs Committee with dedicated staffing and budget.",
            "UNIDO and the Multilateral Fund should ensure that Azerbaijan's HCFC Phase-Out Management Plan (HPMP) Stage II builds directly on enterprise conversion data and lessons from Stage I, avoiding duplication and building on established relationships.",
            "The Ministry of Ecology and Natural Resources should expand the HCFC licensing and quota system to cover all identified importers and distributors, closing regulatory gaps identified during Stage I implementation.",
            "Future Montreal Protocol projects in similar economies should build enterprise verification mechanisms that can be operated by national inspectorates without international consultant support from the start of implementation.",
            "UNIDO should support Azerbaijan in developing refrigerant recovery and recycling infrastructure as a priority in subsequent programme stages, addressing the long-term management of phased-out substances currently lacking end-of-life pathways.",
        ],
    },
    "UNIDO-104112": {
        "project_id":          "104112",
        "title":               "Independent Terminal Evaluation: Promoting the Adaptation and Adoption of RECP (Resource Efficient and Cleaner Production) Through the Establishment and Operation of a Cleaner Production Centre (CPC) in Ukraine",
        "short_title":         "RECP & Cleaner Production Centre – Ukraine",
        "year":                2021,
        "country":             "Ukraine",
        "region":              "Europe and Central Asia",
        "thematic_category":   "Circular Economy / Waste Management",
        "secondary_thematic_area": "Industrial Policy & Competitiveness",
        "report_type":         "Project Evaluation",
        "donor":               "Switzerland (SECO) / Austria",
        "budget_usd":          5_181_779,
        "evaluation_rating":   5.0,
        "overall_rating_label":"Satisfactory",
        "util_rate":           91.4,
        # ── Duration & delay (Oct 2011 – Dec 2020, 50-month overrun) ─────────
        "planned_months":      60,
        "actual_months":       110,
        "delay_months":        50,
        "overrun_pct":         83,
        "has_delay":           True,
        "start_year":          2011,
        "planned_end_year":    2016,
        "actual_end_year":     2020,
        "duration_str":        "Oct 2011 – Dec 2020",
        "delay_causes": (
            "The conflict in eastern Ukraine (2013–14) and annexation of Crimea caused a major "
            "economic shock, slowing industrial production and project uptake across target sectors. "
            "Budget underspending left USD 1.3M available in 2017 (original close year), triggering "
            "36 months of no-cost extension. Organisational delays — creating a second LLC entity "
            "alongside the NGO and extended CTA engagement — compounded the overrun."
        ),
        # ── Year-by-year expenditure (Grants 200001206/07/08 combined) ───────
        "disbursement_by_year": {
            2012: 287_996,
            2013: 331_879,
            2014: 714_686,
            2015: 631_833,
            2016: 333_199,
            2017: 578_989,
            2018: 467_158,
            2019: 266_685,
            2020: 339_958,
        },
        # ── Donor contributions (no standard GEF co-fin table — SECO-funded) ─
        "cofinancing_labels":  ["SECO (Switzerland)", "Austria"],
        "cofinancing_planned": [4_590_000, 591_779],
        "cofinancing_actual":  [4_337_482, 400_000],
        # ── Timeline milestones ───────────────────────────────────────────────
        "timeline_events": [
            {"yr": 2011, "lab": "Project launch",        "type": "plan"},
            {"yr": 2013, "lab": "RECPC NGO established", "type": "actual"},
            {"yr": 2015, "lab": "Mid-Term Review",       "type": "actual"},
            {"yr": 2016, "lab": "RECPC LLC established", "type": "actual"},
            {"yr": 2018, "lab": "1st no-cost extension", "type": "actual"},
            {"yr": 2020, "lab": "Project completion",    "type": "actual"},
        ],
        # ── Gender (Table 7: % female in regional RECP training + RECPC staff)
        "stakeholders_rich": [
            {"name": "RECPC Staff",       "role": "CPC management & assessors",        "pct_women": 40},
            {"name": "Kyiv hub",          "role": "Regional RECP training (2019)",     "pct_women": 33},
            {"name": "Zaporizhzhia hub",  "role": "Regional RECP training (2019)",     "pct_women": 67},
            {"name": "Lviv hub",          "role": "Regional RECP training (2019)",     "pct_women": 56},
            {"name": "Kharkiv hub",       "role": "Regional RECP training (2019)",     "pct_women": 33},
            {"name": "RECP expert pool",  "role": "Certified RECP experts (all years)","pct_women": 35},
        ],
        "gender_rating": "Satisfactory",
        "gender_note": (
            "Gender mainstreaming improved progressively over the project. By close, 5 of 13 RECPC "
            "assessors were women (40%), and 35% of 325 trained RECP experts were female (114 women). "
            "Regional training saw strong upward trends: Lviv rose from 14% to 56% female, "
            "Zaporizhzhia from 56% to 67% (2018–2019). The Centre applied for the Women's Energy "
            "Club of Ukraine gender award in 2020. Rated Satisfactory."
        ),
        # ── Per-criterion DAC ratings (1–6 scale: HS=6, S=5, MS=4, MU=3, U=2, HU=1) ─
        "dac_ratings": {
            "relevance":      5,   # Satisfactory — aligned to national RECP needs and EU integration agenda
            "effectiveness":  5,   # Satisfactory — CPC operational; 325 experts trained; 73 RECP assessments
            "efficiency":     3,   # Moderately Unsatisfactory — 50-month overrun (83%); significant underspend
            "impact":         5,   # Satisfactory — measurable resource savings; institutional embedding achieved
            "sustainability": 4,   # Moderately Satisfactory — CPC financially fragile but institutionally anchored
        },
        "dac_rating_labels": {
            "relevance":      "Satisfactory",
            "effectiveness":  "Satisfactory",
            "efficiency":     "Moderately Unsatisfactory",
            "impact":         "Satisfactory",
            "sustainability": "Moderately Satisfactory",
        },
        "sdgs":                [9, 12, 13, 17],
        "thematic_justification": (
            "The project established and operationalized Ukraine's National Cleaner Production Centre "
            "(CPC) to mainstream Resource-Efficient and Cleaner Production (RECP) methodologies across "
            "Ukrainian industry. RECP is the foundational framework of circular economy thinking in "
            "industrial contexts — reducing material inputs, minimising waste, improving energy "
            "efficiency, and preventing pollution at source."
        ),
        "sdg_justifications": {
            "9":  "By establishing the RECP Centre and providing technical assistance to enterprises across multiple industrial sub-sectors, the project fostered inclusive and sustainable industrialization (SDG 9), promoting innovation in production processes and infrastructure for clean technology adoption.",
            "12": "Resource-Efficient and Cleaner Production (RECP) is the operational methodology for achieving sustainable consumption and production (SDG 12) in industry. The project trained enterprises to reduce raw material consumption, minimize waste, and adopt more sustainable production patterns.",
            "13": "RECP assessments identified energy efficiency improvements and emission reduction opportunities. The adoption of cleaner production technologies directly reduces industrial greenhouse gas emissions, contributing to SDG 13 (Climate Action) at the enterprise and sectoral level.",
            "17": "The project was implemented through a partnership between UNIDO, Switzerland (SECO), and Austrian development cooperation alongside Ukrainian counterparts — a multi-donor, multi-stakeholder approach that exemplifies SDG 17 (Partnerships for the Goals).",
        },
        "lessons_learned": [
            "Embedding a national RECP Centre within an established host institution from project inception significantly increases institutional sustainability and reduces operational fragility compared to standalone project management units.",
            "Integration of RECP methodology into university curricula and vocational training systems creates long-term capacity that persists beyond project timeframes, multiplying impact through trained professionals entering industry.",
            "RECP assessments are most effective when enterprise participation is genuinely voluntary and demand-driven rather than supply-pushed; enterprises that self-select into the programme achieve higher implementation rates of identified improvements.",
            "Developing fee-for-service revenue models for RECP centres from an early stage of project implementation reduces dependence on donor funding and builds a commercially sustainable knowledge services market.",
            "Multi-donor programmes benefit from clear governance arrangements and defined roles between implementing partners at the outset; ambiguity in donor coordination creates delays and inconsistent reporting.",
        ],
        "recommendations": [
            "The Ukrainian RECP Centre should develop and implement a five-year financial sustainability plan that progressively reduces dependence on UNIDO project funding by growing fee-for-service revenues from RECP assessments and training services.",
            "UNIDO should support the RECP Centre in establishing a national RECP policy framework that mandates periodic environmental and resource efficiency audits for enterprises above a defined threshold, creating a sustained demand base for Centre services.",
            "The Ministry of Environmental Protection and Natural Resources of Ukraine should integrate RECP principles and indicators into national industrial development strategies and green economy action plans.",
            "Future RECP Centre projects should include a formal mentoring relationship with an established RECP Centre from a peer country from project inception, accelerating institutional learning and reducing the cost of establishing core competencies.",
            "UNIDO and donors should consider co-financing a green credit facility aligned with the RECP Centre to enable enterprises to finance the implementation of identified RECP improvements, converting assessments into measurable resource savings.",
        ],
    },
    "UNIDO-120323": {
        "project_id":          "GFURU-120323",
        "gef_id":              "4890",
        "title":               "Independent Terminal Evaluation: Towards a Green Economy in Uruguay: Stimulating Sustainable Practices and Low-Emission Technologies in Prioritized Sectors",
        "short_title":         "Green Economy – Uruguay",
        "year":                2021,
        "country":             "Uruguay",
        "region":              "Latin America",
        "thematic_category":   "Climate Action",
        "secondary_thematic_area": "Circular Economy / Waste Management",
        "report_type":         "Project Evaluation",
        "donor":               "GEF",
        "budget_usd":          3_392_727,
        "evaluation_rating":   5.0,
        "overall_rating_label":"Satisfactory",
        "util_rate":           99.4,
        # ── Duration & delay (Dec 2013 – Dec 2020, 37-month overrun) ─────────
        "planned_months":      48,
        "actual_months":       85,
        "delay_months":        37,
        "overrun_pct":         77,
        "has_delay":           True,
        "start_year":          2013,
        "planned_end_year":    2017,
        "actual_end_year":     2020,
        "duration_str":        "Dec 2013 – Dec 2020",
        "delay_causes": (
            "The original 48-month timeline was unrealistic given the project's technical and "
            "institutional complexity. Significant disbursement lag in the first three years (only "
            "USD 99,511 of USD 446,694 planned in 2014) reflected procedural shortcomings and "
            "limited proactivity of monitoring structures. INC's flagship biodigester pilot was "
            "discarded after pre-feasibility and ALUR delivered only 26% of its planned investments, "
            "requiring major restructuring of pilot activities."
        ),
        # ── GEF expenditure by component (year-by-year chart unreadable in PDF)
        "disbursement_by_year": {
            "C1: Policy": 677_508,
            "C2: Knowledge": 394_934,
            "C3: Pilots": 1_766_205,
            "C4: Capacity": 469_268,
            "C5: M&E": 48_594,
        },
        # ── Co-financing by source: planned vs actual (Table 8) ───────────────
        "cofinancing_labels":  ["INC", "ALUR", "Estancias del Lago", "MGAP", "MIEM", "UNIDO"],
        "cofinancing_planned": [13_000_000, 7_400_000, 10_000_000, 300_000, 1_300_000, 110_000],
        "cofinancing_actual":  [0, 1_924_000, 10_000_000, 300_000, 1_300_000, 110_000],
        # ── Timeline milestones ───────────────────────────────────────────────
        "timeline_events": [
            {"yr": 2013, "lab": "Project start",          "type": "plan"},
            {"yr": 2016, "lab": "Mid-Term Review",        "type": "actual"},
            {"yr": 2017, "lab": "Planned completion",     "type": "plan"},
            {"yr": 2017, "lab": "Biogas standard adopted","type": "actual"},
            {"yr": 2019, "lab": "Circular Econ. Network", "type": "actual"},
            {"yr": 2020, "lab": "Actual completion",      "type": "actual"},
        ],
        # ── Gender indicators (Annex 6, PIR 2020 — targets all exceeded) ─────
        "stakeholders_rich": [
            {"name": "DNE/DINAMA/MGAP WG", "role": "Policy working group",        "pct_women": 60},
            {"name": "Knowledge Network",   "role": "Knowledge mgmt platform",     "pct_women": 48},
            {"name": "Workshop attendees",  "role": "Dissemination workshops",     "pct_women": 67},
            {"name": "University WGs",      "role": "Academia–industry liaison",   "pct_women": 67},
        ],
        "gender_rating": "Satisfactory",
        "gender_note": (
            "All gender targets were achieved and in most cases exceeded. The inter-institutional "
            "policy working group reached 60% female (target: 30%). The knowledge network was 48% "
            "female; dissemination workshops 67%; university working groups 67%. A gender action "
            "plan aligned with UNIDO's 2016 Gender Equality strategy was implemented from Year 2. "
            "Rated Satisfactory (Score 5)."
        ),
        # ── Per-criterion DAC ratings (1–6 scale: HS=6, S=5, MS=4, MU=3, U=2, HU=1) ─
        "dac_ratings": {
            "relevance":      6,   # Highly Satisfactory — directly aligned to Uruguay NDC and green economy strategy
            "effectiveness":  5,   # Satisfactory — all output targets met; policy frameworks adopted
            "efficiency":     3,   # Moderately Unsatisfactory — 37-month overrun; slow early disbursement
            "impact":         5,   # Satisfactory — GHG reductions demonstrated; green economy networks established
            "sustainability": 5,   # Satisfactory — institutionalised in MoE; financial mechanisms embedded
        },
        "dac_rating_labels": {
            "relevance":      "Highly Satisfactory",
            "effectiveness":  "Satisfactory",
            "efficiency":     "Moderately Unsatisfactory",
            "impact":         "Satisfactory",
            "sustainability": "Satisfactory",
        },
        "sdgs":                [7, 9, 12, 13, 17],
        "thematic_justification": (
            "The project's primary objective was to stimulate the adoption of sustainable practices "
            "and low-emission technologies in key Uruguayan industrial and agricultural sectors, "
            "with a direct focus on reducing greenhouse gas emissions and transitioning towards a "
            "green economy. This encompasses renewable energy deployment, energy efficiency in "
            "industry, and low-carbon production methods — placing it in the Climate Action theme."
        ),
        "sdg_justifications": {
            "7":  "The project supported deployment of renewable energy technologies — including solar, wind, and biomass — in prioritized sectors in Uruguay, contributing to SDG 7 (Affordable and Clean Energy) by increasing the share of renewables and reducing fossil fuel dependence.",
            "9":  "By providing technical assistance and financing to enterprises adopting low-emission technologies and cleaner production methods, the project promoted sustainable industrialization and innovation (SDG 9) across Uruguay's priority industrial and agro-industrial sectors.",
            "12": "The project stimulated adoption of sustainable production practices — including resource efficiency improvements, waste reduction, and clean technology adoption — advancing responsible consumption and production patterns (SDG 12) at enterprise and sectoral level.",
            "13": "Low-emission technology deployment in industry directly reduces Uruguay's greenhouse gas emissions. The project contributes to SDG 13 (Climate Action) by demonstrating viable industrial decarbonization pathways and providing evidence for national climate policy.",
            "17": "Funded by GEF and implemented through UNIDO in partnership with Uruguay's Ministry of Environment and Ministry of Industry — with multiple private sector actors — this project exemplifies the multi-stakeholder partnerships of SDG 17 (Partnerships for the Goals).",
        },
        "lessons_learned": [
            "Flexible programme design that allows resource reallocation between technology tracks — based on demonstrated uptake and market readiness — significantly outperforms rigid activity plans in complex multi-technology green economy programmes.",
            "Establishing financial mechanisms (green credit lines, revolving funds) early in project implementation rather than at mid-term is critical, as financial instrument setup requires longer lead times than physical technology demonstrations.",
            "Robust baseline data collection on energy consumption and emissions at the enterprise level is essential for demonstrating project impact; without reliable baselines, attribution of GHG reductions to project interventions is contested.",
            "Strong alignment between project objectives and national policy commitments — such as Nationally Determined Contributions — creates political ownership that sustains programme momentum through administrative transitions.",
            "Private sector co-financing commitments in green technology projects are most successfully mobilized when complemented by risk-reduction instruments; grants alone do not address the financial barriers to first-mover technology adoption.",
        ],
        "recommendations": [
            "The Government of Uruguay should embed the green economy financial mechanisms established under this project — particularly the green credit line — within permanent public financial institution operations to ensure their continuity beyond GEF funding.",
            "UNIDO should support Uruguay in developing a national green economy monitoring framework with standardised enterprise-level energy and emissions reporting, enabling accurate tracking of NDC contributions from the industrial sector.",
            "Future GEF green economy programmes in upper-middle-income countries should prioritise catalytic financial instruments over grant-based technology subsidies, recognising that the primary barrier is financial risk rather than technology availability.",
            "The Ministry of Industry of Uruguay should use evidence from this project's technology demonstrations to update sectoral energy efficiency standards and incentive frameworks, institutionalising the transition to low-emission technologies.",
            "UNIDO and GEF should explore a programmatic approach to green economy support in Latin America that enables Uruguay and peer countries to share technology assessment data, financing instruments, and market development lessons across country boundaries.",
        ],
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────────────────

st.markdown("""
<style>
  :root {
    --blue: #009EDB; --dark: #1a1a2e; --bg: #f5f7fa;
    --muted: #6b7280; --border: #e5e7eb;
    --danger: #ef4444; --success: #22c55e; --amber: #f59e0b;
  }
  #MainMenu, footer, header { visibility: hidden; }
  .main .block-container {
    padding-top: 1rem; padding-bottom: 2rem;
    max-width: 100% !important; padding-left: 1.5rem; padding-right: 1.5rem;
  }
  /* Remove extra top padding Streamlit adds above columns */
  div[data-testid="stVerticalBlock"] > div { gap: 0.4rem; }
  /* Tighten column gaps */
  div[data-testid="stHorizontalBlock"] { gap: 0.5rem; }

  .unido-header {
    background: white; border-bottom: 2px solid var(--blue);
    padding: 0.8rem 1.5rem; margin: -1.2rem -1rem 1.2rem -1rem;
    display: flex; align-items: center; justify-content: space-between;
  }
  .unido-logo { font-size: 1.3rem; font-weight: 700; color: var(--blue); }
  .unido-sub  { font-size: 0.85rem; color: var(--muted); }
  .user-chip  {
    background: var(--bg); border: 1px solid var(--border);
    border-radius: 999px; padding: 0.3rem 0.9rem;
    font-size: 0.82rem; color: var(--dark);
  }

  /* Login */
  .login-wrap { max-width: 400px; margin: 3rem auto; }
  .login-logo { text-align:center; margin-bottom: 1.5rem; }
  .login-logo .big { font-size: 2rem; font-weight: 700; color: var(--blue); }
  .login-logo .sub { color: var(--muted); font-size: 0.85rem; }

  /* Report cards */
  .report-card {
    background: white; border: 1px solid var(--border);
    border-radius: 10px; padding: 1.1rem 1.3rem; margin-bottom: 0.5rem;
    box-shadow: 0 2px 8px rgba(0,0,0,0.07);
    border-top: 3px solid var(--blue);
    transition: box-shadow 0.15s ease;
  }
  .report-card:hover { box-shadow: 0 4px 16px rgba(0,0,0,0.12); }
  .report-title { font-weight: 700; color: var(--dark); font-size: 0.92rem; line-height:1.45; margin-bottom:0.3rem; }
  .report-meta  { color: var(--muted); font-size: 0.76rem; margin: 0.1rem 0 0.35rem; }
  .tag {
    display: inline-block; border-radius: 999px; font-size: 0.72rem;
    font-weight: 600; padding: 0.15rem 0.55rem; margin: 2px;
  }
  .tag-blue  { background: #e0f2fe; color: #0369a1; }
  .tag-green { background: #dcfce7; color: #166534; }
  .tag-amber { background: #fef3c7; color: #92400e; }
  .tag-gray  { background: #f3f4f6; color: #374151; }
  .tag-purple { background: #f3e8ff; color: #6d28d9; }

  /* Rating badge */
  .rating-pill {
    display:inline-flex; align-items:center; gap:4px;
    border-radius:999px; font-size:0.72rem; font-weight:700;
    padding:0.18rem 0.6rem; margin:2px;
  }
  .rating-high   { background:#dcfce7; color:#166534; }
  .rating-mid    { background:#fef3c7; color:#92400e; }
  .rating-low    { background:#fee2e2; color:#991b1b; }
  .rating-none   { background:#f3f4f6; color:#6b7280; }

  /* Passage cards */
  .passage-card {
    background: var(--bg); border-left: 3px solid var(--blue);
    border-radius: 0 6px 6px 0; padding: 0.8rem 1rem;
    margin-bottom: 0.5rem; font-size: 0.87rem;
  }
  .passage-title { font-weight: 600; color: var(--dark); font-size: 0.85rem; }
  .passage-meta  { color: var(--muted); font-size: 0.77rem; margin-bottom: 0.35rem; }
  .passage-text  { color: #374151; line-height: 1.55; }
  .score-pill {
    display: inline-block; background: #e0f2fe; color: #0369a1;
    font-size: 0.72rem; font-weight: 600; border-radius: 999px;
    padding: 0.1rem 0.45rem; margin-left: 0.4rem;
  }

  /* DAC evidence */
  .dac-card {
    background: white; border: 1px solid var(--border);
    border-radius: 6px; padding: 0.9rem 1.1rem; margin-bottom: 0.5rem;
  }
  .dac-quote {
    border-left: 3px solid var(--blue); padding-left: 0.8rem;
    color: #374151; font-size: 0.86rem; line-height: 1.55;
    font-style: italic;
  }

  /* Stat cards */
  .stat-card {
    background: white; border: 1px solid var(--border); border-radius: 8px;
    padding: 1.2rem; text-align: center; box-shadow: 0 1px 3px rgba(0,0,0,0.05);
  }
  .stat-num  { font-size: 2rem; font-weight: 700; color: var(--blue); }
  .stat-lbl  { color: var(--muted); font-size: 0.8rem; margin-top: 0.2rem; }

  /* Sidebar */
  [data-testid="stSidebar"] { background: white; border-right: 1px solid var(--border); }
  [data-testid="stSidebar"] .stMarkdown h3 {
    color: var(--dark); font-size: 0.78rem; text-transform: uppercase;
    letter-spacing: 0.05em; font-weight: 600; margin-top: 1rem;
  }

  /* User mgmt */
  .role-badge { display:inline-block; padding:.15rem .6rem; border-radius:999px; font-size:.72rem; font-weight:600; }
  .role-admin { background:#fef3c7; color:#92400e; }
  .role-user  { background:#e0f2fe; color:#0369a1; }
  .dot { display:inline-block; width:8px; height:8px; border-radius:50%; margin-right:5px; }
  .dot-green { background:var(--success); }
  .dot-red   { background:var(--danger); }
  .dot-amber { background:var(--amber); }

  [data-testid="baseButton-primary"] {
    background-color: var(--blue) !important;
    border-color: var(--blue) !important;
  }
  .stButton > button {
    border-radius: 6px;
    font-size: 0.8rem;
    padding: 0.3rem 0.85rem;
  }

  /* Tight button row — correct Streamlit selectors */
  div[data-testid="stHorizontalBlock"]:has([data-testid="stButton"]) {
    gap: 6px !important;
  }
  div[data-testid="stHorizontalBlock"]:has([data-testid="stButton"]) > div[data-testid="column"] {
    padding-left: 0 !important;
    padding-right: 0 !important;
    min-width: 0 !important;
    flex: 0 0 auto !important;
    width: auto !important;
  }
  div[data-testid="stHorizontalBlock"]:has([data-testid="stButton"]) > div[data-testid="column"]:last-child {
    flex: 1 1 auto !important;
  }
  div[data-testid="stHorizontalBlock"]:has([data-testid="stButton"]) .stButton > button,
  div[data-testid="stHorizontalBlock"]:has([data-testid="stButton"]) .stDownloadButton > button {
    white-space: nowrap !important;
    width: auto !important;
  }

  /* Card action buttons */
  .card-actions {
    display: flex; gap: 8px; margin-top: 0.5rem; margin-bottom: 0.2rem;
  }
  .action-btn {
    flex: 1; padding: 0.38rem 0.7rem; border-radius: 6px; font-size: 0.78rem;
    font-weight: 600; cursor: pointer; border: none; text-align: center;
    font-family: inherit;
  }
  .action-btn-outline {
    background: white; color: var(--dark); border: 1.5px solid var(--border);
  }
  .action-btn-outline:hover { border-color: var(--blue); color: var(--blue); }
  .action-btn-primary { background: var(--blue); color: white; }
  .action-btn-primary:hover { background: #007ab8; }
  .action-btn-ghost  { background: #f3f4f6; color: var(--dark); border: 1.5px solid transparent; }
  .action-btn-ghost:hover { background: #e5e7eb; }

  /* Pilot banner */
  .pilot-banner {
    background: #eff6ff; border: 1px solid #bfdbfe; border-radius: 8px;
    padding: 0.6rem 1rem; margin-bottom: 1rem; font-size: 0.82rem; color: #1e40af;
  }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# Session state
# ─────────────────────────────────────────────────────────────────────────────

import time as _time

for k, v in {
    "session_token": "no-auth-bypass", "username": "guest",
    "display_name": "Guest", "role": "admin",
    "all_reports": None,
    "lessons_cache": {},
    "synth_history": [],
    "backend_healthy": None,
    "_page_load": 0,
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# Reset report cache when app restarts (new deployment / new session)
_deploy_ts = int(os.path.getmtime(__file__))
if st.session_state.get("_deploy_ts") != _deploy_ts:
    st.session_state["_deploy_ts"] = _deploy_ts
    st.session_state["all_reports"] = None  # force refresh on new deploy

# ─────────────────────────────────────────────────────────────────────────────
# API helpers
# ─────────────────────────────────────────────────────────────────────────────

def auth_headers() -> dict:
    return {"X-Session-Token": st.session_state.session_token or ""}

def display_title(rep: dict) -> str:
    """Return a clean, concise display title for dropdowns and labels.
    Uses short_title if available, otherwise strips the common ITE preamble."""
    if rep.get("short_title"):
        return rep["short_title"]
    title = rep.get("title", "") or ""
    # Strip common preamble variations
    for prefix in [
        "Independent Terminal Evaluation: The Project ",
        "Independent Terminal Evaluation: ",
        "Terminal Evaluation: ",
    ]:
        if title.startswith(prefix):
            title = title[len(prefix):].strip().strip('"')
            break
    return title or rep.get("report_id", "Unknown")


def api(method: str, path: str, **kw) -> httpx.Response:
    h = kw.pop("headers", {})
    h.update(auth_headers())
    return httpx.request(method, f"{BACKEND_URL}{path}", headers=h, timeout=45, **kw)

def load_reports(force: bool = False):
    if st.session_state.all_reports is None or force:
        # Pull from both sources and merge:
        # metadata.yaml → ratings, full titles, donor, budget
        # Qdrant list   → SDGs (not stored in metadata.yaml)
        meta_reports = load_metadata_reports()
        qdrant_reports: list[dict] = []
        try:
            r = api("GET", "/api/v1/reports/list")
            if r.status_code == 200:
                qdrant_reports = r.json().get("reports", [])
        except Exception:
            pass

        # Build SDG lookup from Qdrant
        sdg_by_rid = {r["report_id"]: r.get("sdgs") or [] for r in qdrant_reports}

        if meta_reports:
            # Use metadata.yaml as base; patch SDGs from Qdrant
            merged = []
            for rep in meta_reports:
                rid = rep.get("report_id", "")
                rep["sdgs"] = sdg_by_rid.get(rid) or rep.get("sdgs") or []
                merged.append(rep)
        else:
            # Fall back to Qdrant only
            merged = qdrant_reports

        # Apply pilot metadata overrides — guarantees correct titles/ratings
        # regardless of what Qdrant or the backend returns
        for rep in merged:
            rid = rep.get("report_id", "")
            if rid in PILOT_METADATA:
                rep.update(PILOT_METADATA[rid])

        # Always ensure the 4 pilot reports are present even if backend is unreachable
        existing_ids = {r.get("report_id") for r in merged}
        for rid, meta in PILOT_METADATA.items():
            if rid not in existing_ids:
                merged.append({"report_id": rid, **meta})

        st.session_state.all_reports = sorted(
            merged, key=lambda r: r.get("year") or 0, reverse=True
        )

def load_lessons(report_ids: list[str]) -> list[dict]:
    try:
        r = api("GET", "/api/v1/lessons", params={"report_ids": ",".join(report_ids)})
        return r.json().get("reports", []) if r.status_code == 200 else []
    except Exception:
        return []

def _load_sections_local(report_id: str) -> dict:
    """Load extracted sections directly from local JSON file (no backend needed).
    Works on Streamlit Cloud since data/ is in the git repo."""
    import pathlib
    p = pathlib.Path(__file__).parent.parent / "data" / "extracted_sections" / f"{report_id}.json"
    if p.exists():
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def load_sections(report_id: str) -> dict | None:
    """Load pre-extracted PDF sections — tries local file first, then backend API."""
    # Try local file first (faster, works on Streamlit Cloud)
    local = _load_sections_local(report_id)
    if local:
        return local
    # Fallback to backend API
    try:
        r = api("GET", f"/api/v1/reports/{report_id}/sections")
        return r.json() if r.status_code == 200 else None
    except Exception:
        return None

def load_metadata_reports() -> list[dict]:
    """Load full metadata (incl. ratings) from metadata.yaml via backend."""
    try:
        r = api("GET", "/api/v1/metadata")
        return r.json().get("reports", []) if r.status_code == 200 else []
    except Exception:
        return []

def do_login(username, password):
    try:
        r = httpx.post(f"{BACKEND_URL}/api/v1/auth/login",
                       json={"username": username, "password": password}, timeout=10)
        if r.status_code == 200:
            d = r.json()
            st.session_state.update({
                "session_token": d["session_token"],
                "username":      d["username"],
                "display_name":  d["display_name"],
                "role":          d["role"],
            })
            return True, ""
        return False, r.json().get("detail", "Login failed.")
    except httpx.ConnectError:
        return False, f"Cannot reach backend at {BACKEND_URL}. Is it running?"
    except Exception as e:
        return False, str(e)

def do_logout():
    try:
        api("POST", "/api/v1/auth/logout")
    except Exception:
        pass
    for k in ["session_token", "username", "display_name", "role",
               "all_reports", "lessons_cache", "synth_history"]:
        st.session_state[k] = None if k not in ("lessons_cache", "synth_history") \
                                else ({} if k == "lessons_cache" else [])
    st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# Excel export helper
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────────────────────────────────────────────────────────────────────
# Text cleaning — strip PDF extraction artefacts and split into items
# ─────────────────────────────────────────────────────────────────────────────

_TOC_LINE  = re.compile(r'^.{3,60}[.\s]{3,}\d{1,3}\s*$')
_PAGE_NUM  = re.compile(r'^\s*\d{1,3}\s*$')

# Patterns that indicate questionnaire / ToR / garbage lines
_QUESTION_STARTS = (
    "what ", "how ", "did ", "do you", "are you", "were you", "please ",
    "who ", "when ", "why ", "which ", "have you", "would you", "could you",
    "is there", "was there", "can you",
)
_GARBAGE_PATTERNS = (
    "terms of reference", "table of contents", "list of figures",
    "list of tables", "abbreviations", "acronyms", "questionnaire",
    "interview guide", "contact person", "please email",
    "name of your", "your position", "name of company",
    "highly unsatisfactory", "highly satisfactory", "vienna international",
    "wagramerstr", "minimum organizational", "advanced degree",
    "key informant interview", "questions for key", "criteria question",
    "figures and tables", "project factsheet", "acknowledgements",
    "actual project start", "planned project completion", "project duration",
    "gef ceo endorsement", "pad issuance", "first august", "ful issuance",
    "soalan temubual", "membangunkan dapatan", "cadangan untuk",
    # Intro sentences that introduce but are not lessons/recs themselves
    "the following lessons are", "the following recommendations are",
    "the following lessons were", "the following recommendations were",
    "lessons are proposed for consideration", "are proposed for consideration",
    "following key lessons", "lessons identified in this evaluation",
)

# Words that signal a fragment starting mid-sentence (not a genuine lesson)
_FRAGMENT_STARTS = (
    "of ", "to ", "and ", "in ", "with ", "by ", "for ", "from ", "at ",
    "on ", "as ", "but ", "or ", "nor ", "so ", "yet ", "both ", "either ",
    "neither ", "not ", "only ", "also ", "thus ", "hence ", "however ",
    "moreover ", "furthermore ", "therefore ", "additionally ", "similarly ",
    "consequently ", "nevertheless ", "nonetheless ",
)


def _is_garbage_chunk(text: str) -> bool:
    """Return True if this chunk is mostly questionnaire / ToR / non-lesson content."""
    if not text:
        return True
    lower = text.lower()
    for pat in _GARBAGE_PATTERNS:
        if pat in lower:
            return True
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    if not lines:
        return True
    # Reject if >35% of lines are questions
    q_count = sum(1 for l in lines if l.endswith("?") or
                  any(l.lower().startswith(s) for s in _QUESTION_STARTS))
    if q_count / len(lines) > 0.35:
        return True
    # Reject if chunk looks like a project data table (many short date-like lines)
    date_re = re.compile(r'\b(january|february|march|april|may|june|july|august|'
                         r'september|october|november|december|\d{4})\b', re.I)
    date_lines = sum(1 for l in lines if date_re.search(l) and len(l) < 80)
    if len(lines) >= 3 and date_lines / len(lines) > 0.5:
        return True
    return False


def _is_quality_item(text: str) -> bool:
    """Return True if this extracted item looks like a genuine lesson/recommendation."""
    t = text.strip()
    if len(t) < 50:
        return False
    if t.endswith("?") or t.endswith(":"):
        return False
    lower = t.lower()
    # Reject questionnaire-style starts
    for s in _QUESTION_STARTS:
        if lower.startswith(s):
            return False
    # Reject obvious garbage patterns
    for pat in _GARBAGE_PATTERNS:
        if pat in lower:
            return False
    # Must have at least 8 words
    if len(t.split()) < 8:
        return False
    # Reject mid-sentence fragments (start with lowercase continuation word)
    first_char = t[0]
    if first_char.islower():
        return False
    # Reject if starts with a fragment connector word (even if capitalised after stripping)
    for frag in _FRAGMENT_STARTS:
        if lower.startswith(frag):
            return False
    # Reject abbreviation list entries (e.g. "RBM Results-based Management")
    words = t.split()
    if len(words) <= 6 and words[0].isupper() and len(words[0]) <= 6:
        return False
    # Reject items that are clearly table/figure captions
    if re.match(r'^(table|figure|annex|appendix)\s+\d', lower):
        return False
    # Detect garbled PDF table text: institution name repeated 2+ times in first 160 chars
    # e.g. "Recommendation 2: UNIDO and GEF ... UNIDO During the ..."
    head = lower[:160]
    for inst in ("unido", "gef ", "undp", "unep"):
        if head.count(inst) >= 2:
            return False
    # Reject intro sentences ending with colon (they introduce a list, not a lesson)
    if re.search(r'(following|consideration|below|noted|identified)\s*:\s*$', lower):
        return False
    # Must contain at least one verb-like structure
    if not re.search(r'\b(should|must|need|ensure|improve|strengthen|consider|'
                     r'increase|reduce|support|develop|establish|provide|include|'
                     r'require|recommend|is|are|was|were|will|would|can|could|'
                     r'has|have|had|been|be|demonstrated|showed|found|proved|'
                     r'contributed|resulted|enabled|allowed)\b', lower):
        return False
    return True


def _clean_raw_text(text: str) -> str:
    """Remove table-of-contents lines, page numbers, and other PDF artefacts."""
    if not text:
        return ""
    lines = []
    for line in text.split("\n"):
        stripped = line.strip()
        if not stripped:
            lines.append("")
            continue
        if _TOC_LINE.match(stripped):
            continue
        if _PAGE_NUM.match(stripped):
            continue
        lines.append(line)
    cleaned = "\n".join(lines)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def extract_items(chunk_text: str, max_items: int = 8) -> list[str]:
    # Reject entire chunk if it is questionnaire/ToR garbage
    if _is_garbage_chunk(chunk_text):
        return []

    text = _clean_raw_text(chunk_text)
    if not text:
        return []

    def cap_text(s):
        s = s.strip()
        count = 0
        for i in range(len(s) - 2):
            if s[i] in ".!?" and s[i + 1] == " " and s[i + 2].isupper():
                count += 1
                if count >= 2:
                    s = s[: i + 1]
                    break
        return s[:320]

    items = []
    current = []
    in_item = False

    for line in text.split("\n"):
        stripped = line.strip()
        if not stripped:
            continue
        j = 0
        while j < 2 and j < len(stripped) and stripped[j].isdigit():
            j += 1
        is_new = (
            j > 0
            and j < len(stripped)
            and stripped[j] in ".)"
            and j + 1 < len(stripped)
            and stripped[j + 1] == " "
        )
        if is_new:
            if current:
                txt = cap_text(" ".join(current))
                if _is_quality_item(txt):
                    items.append(txt)
                if len(items) >= max_items:
                    break
            current = [stripped[j + 2:]]
            in_item = True
        elif in_item:
            current.append(stripped)

    if current and len(items) < max_items:
        txt = cap_text(" ".join(current))
        if _is_quality_item(txt):
            items.append(txt)

    if items:
        return items

    # Fallback: return the chunk as a single item if it passes quality check
    result = cap_text(text)
    return [result] if _is_quality_item(result) else []



def make_excel_sections(reports_meta: list[dict], sections_by_id: dict) -> bytes:
    """Comprehensive multi-sheet Excel export with all view detail fields."""
    import pathlib
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    SDG_NAMES_LOCAL = {
        1:"No Poverty",2:"Zero Hunger",3:"Good Health",4:"Quality Education",
        5:"Gender Equality",6:"Clean Water",7:"Affordable Clean Energy",
        8:"Decent Work",9:"Industry Innovation",10:"Reduced Inequalities",
        11:"Sustainable Cities",12:"Responsible Consumption",13:"Climate Action",
        14:"Life Below Water",15:"Life on Land",16:"Peace Justice",17:"Partnerships",
    }

    HEADER_FILL  = PatternFill("solid", fgColor="003DA5")
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL     = PatternFill("solid", fgColor="EFF6FF")
    WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
    TOP_ALIGN    = Alignment(vertical="top")

    def _style_sheet(ws, col_widths: list[int]):
        """Apply header styling and column widths."""
        for cell in ws[1]:
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        # Alternate row shading + wrap
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = ALT_FILL if row_idx % 2 == 0 else None
            row_h = 15
            for cell in row:
                if fill:
                    cell.fill = fill
                cell.alignment = WRAP_ALIGN
                val_len = len(str(cell.value or ""))
                row_h = max(row_h, min(15 * max(1, val_len // 60), 120))
            ws.row_dimensions[row_idx].height = row_h

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Sheet 1: Metadata ────────────────────────────────────────────────
        meta_rows = []
        for rep in reports_meta:
            rid  = rep.get("report_id", "")
            ai   = _load_ai_extraction(rid)
            ctx  = ai.get("context", {})
            rating = ctx.get("evaluation_rating") or rep.get("evaluation_rating")
            sdgs_raw = ai.get("sdg_mapping", {})
            sdg_nums = sorted(int(k) for k in sdgs_raw.keys() if str(k).isdigit()) if sdgs_raw else (rep.get("sdgs") or [])
            meta_rows.append({
                "Report ID":               rid,
                "Title":                   ctx.get("title") or rep.get("title", ""),
                "Year":                    ctx.get("year") or rep.get("year", ""),
                "Country":                 ctx.get("country") or rep.get("country", ""),
                "Region":                  ctx.get("region") or rep.get("region", ""),
                "Report Type":             ctx.get("report_type") or rep.get("report_type", ""),
                "Primary Thematic Area":   ai.get("primary_thematic_area") or rep.get("thematic_category", ""),
                "Secondary Thematic Area": ai.get("secondary_thematic_area", ""),
                "Thematic Justification":  ai.get("thematic_justification", ""),
                "Evaluation Rating":       f"{float(rating):.1f}/6" if rating else "",
                "Donor":                   ctx.get("donor") or rep.get("donor", ""),
                "Project ID":              ctx.get("project_id") or rep.get("project_id", ""),
                "Budget (USD)":            ctx.get("budget_usd") or rep.get("budget_usd", ""),
                "SDGs":                    ", ".join(f"SDG {n}" for n in sdg_nums),
            })
        df_meta = pd.DataFrame(meta_rows)
        df_meta.to_excel(writer, sheet_name="Metadata", index=False)
        ws = writer.sheets["Metadata"]
        _style_sheet(ws, [14,55,6,18,14,22,28,28,55,16,20,14,14,30])

        # ── Sheet 2: Executive Summary ───────────────────────────────────────
        summ_rows = []
        for rep in reports_meta:
            rid = rep.get("report_id", "")
            ai  = _load_ai_extraction(rid)
            ctx = ai.get("context", {})
            summ_rows.append({
                "Report ID": rid,
                "Title":     ctx.get("title") or rep.get("title", ""),
                "Year":      ctx.get("year") or rep.get("year", ""),
                "Country":   ctx.get("country") or rep.get("country", ""),
                "Executive Summary": ai.get("executive_summary") or rep.get("executive_summary", ""),
            })
        df_summ = pd.DataFrame(summ_rows)
        df_summ.to_excel(writer, sheet_name="Executive Summary", index=False)
        ws = writer.sheets["Executive Summary"]
        _style_sheet(ws, [14,55,6,18,100])

        # ── Sheet 3: Findings & Conclusions ──────────────────────────────────
        fc_rows = []
        for rep in reports_meta:
            rid = rep.get("report_id", "")
            ai  = _load_ai_extraction(rid)
            ctx = ai.get("context", {})
            sec_data = sections_by_id.get(rid, {})
            secs = (sec_data.get("sections", {}) if sec_data else {})
            findings    = secs.get("findings") or secs.get("results") or ""
            conclusions = secs.get("conclusions") or ""
            fc_rows.append({
                "Report ID":   rid,
                "Title":       ctx.get("title") or rep.get("title", ""),
                "Year":        ctx.get("year") or rep.get("year", ""),
                "Country":     ctx.get("country") or rep.get("country", ""),
                "Findings":    findings,
                "Conclusions": conclusions,
            })
        df_fc = pd.DataFrame(fc_rows) if fc_rows else pd.DataFrame(
            columns=["Report ID","Title","Year","Country","Findings","Conclusions"])
        df_fc.to_excel(writer, sheet_name="Findings & Conclusions", index=False)
        ws = writer.sheets["Findings & Conclusions"]
        _style_sheet(ws, [14,45,6,18,90,90])

        # ── Sheet 4: Lessons Learned ─────────────────────────────────────────
        ll_rows = []
        for rep in reports_meta:
            rid = rep.get("report_id", "")
            ai  = _load_ai_extraction(rid)
            ctx = ai.get("context", {})
            lessons = ai.get("lessons_learned") or rep.get("lessons_learned") or []
            if not lessons:
                # fallback: raw sections text
                sec_data = sections_by_id.get(rid, {})
                raw = (sec_data.get("sections", {}) if sec_data else {}).get("lessons_learned", "")
                lessons = [raw] if raw else []
            for i, lesson in enumerate(lessons, 1):
                ll_rows.append({
                    "Report ID": rid,
                    "Title":     ctx.get("title") or rep.get("title", ""),
                    "Year":      ctx.get("year") or rep.get("year", ""),
                    "Country":   ctx.get("country") or rep.get("country", ""),
                    "Primary Thematic Area": ai.get("primary_thematic_area") or rep.get("thematic_category", ""),
                    "#":         i,
                    "Lesson Learned": lesson,
                })
        df_ll = pd.DataFrame(ll_rows) if ll_rows else pd.DataFrame(
            columns=["Report ID","Title","Year","Country","Primary Thematic Area","#","Lesson Learned"])
        df_ll.to_excel(writer, sheet_name="Lessons Learned", index=False)
        ws = writer.sheets["Lessons Learned"]
        _style_sheet(ws, [14,45,6,18,28,4,100])

        # ── Sheet 4: Recommendations ─────────────────────────────────────────
        rec_rows = []
        for rep in reports_meta:
            rid = rep.get("report_id", "")
            ai  = _load_ai_extraction(rid)
            ctx = ai.get("context", {})
            recs = ai.get("recommendations") or rep.get("recommendations") or []
            if not recs:
                sec_data = sections_by_id.get(rid, {})
                raw = (sec_data.get("sections", {}) if sec_data else {}).get("recommendations", "")
                recs = [raw] if raw else []
            for i, rec in enumerate(recs, 1):
                rec_rows.append({
                    "Report ID": rid,
                    "Title":     ctx.get("title") or rep.get("title", ""),
                    "Year":      ctx.get("year") or rep.get("year", ""),
                    "Country":   ctx.get("country") or rep.get("country", ""),
                    "Primary Thematic Area": ai.get("primary_thematic_area") or rep.get("thematic_category", ""),
                    "#":         i,
                    "Recommendation": rec,
                })
        df_rec = pd.DataFrame(rec_rows) if rec_rows else pd.DataFrame(
            columns=["Report ID","Title","Year","Country","Primary Thematic Area","#","Recommendation"])
        df_rec.to_excel(writer, sheet_name="Recommendations", index=False)
        ws = writer.sheets["Recommendations"]
        _style_sheet(ws, [14,45,6,18,28,4,100])

        # ── Sheet 5: SDG Mapping ─────────────────────────────────────────────
        sdg_rows = []
        for rep in reports_meta:
            rid = rep.get("report_id", "")
            ai  = _load_ai_extraction(rid)
            ctx = ai.get("context", {})
            sdg_map = ai.get("sdg_mapping", {})
            if not sdg_map:
                # build basic rows from sdgs list without justifications
                for n in (rep.get("sdgs") or []):
                    try:
                        n = int(n)
                    except Exception:
                        continue
                    sdg_rows.append({
                        "Report ID": rid,
                        "Title":     ctx.get("title") or rep.get("title", ""),
                        "Year":      ctx.get("year") or rep.get("year", ""),
                        "Country":   ctx.get("country") or rep.get("country", ""),
                        "SDG #":     n,
                        "SDG Name":  SDG_NAMES_LOCAL.get(n, ""),
                        "Justification": "",
                    })
            else:
                for k, justification in sdg_map.items():
                    try:
                        n = int(k)
                    except Exception:
                        continue
                    sdg_rows.append({
                        "Report ID": rid,
                        "Title":     ctx.get("title") or rep.get("title", ""),
                        "Year":      ctx.get("year") or rep.get("year", ""),
                        "Country":   ctx.get("country") or rep.get("country", ""),
                        "SDG #":     n,
                        "SDG Name":  SDG_NAMES_LOCAL.get(n, ""),
                        "Justification": justification,
                    })
        df_sdg = pd.DataFrame(sdg_rows) if sdg_rows else pd.DataFrame(
            columns=["Report ID","Title","Year","Country","SDG #","SDG Name","Justification"])
        if not df_sdg.empty and "SDG #" in df_sdg.columns:
            df_sdg = df_sdg.sort_values(["Report ID","SDG #"])
        df_sdg.to_excel(writer, sheet_name="SDG Mapping", index=False)
        ws = writer.sheets["SDG Mapping"]
        _style_sheet(ws, [14,45,6,18,6,26,100])

    return buf.getvalue()


def make_excel(reports_data: list[dict]) -> bytes:
    # Build Excel: one row per extracted lesson or recommendation item.
    rows_l, rows_r = [], []
    for rep in reports_data:
        meta = {
            "Report Title": rep.get("title", ""),
            "Year":         rep.get("year", ""),
            "Country":      rep.get("country", ""),
            "Region":       rep.get("region", ""),
            "Thematic Area":rep.get("thematic_category", ""),
        }
        for chunk in rep.get("lessons_learned", []):
            for item in extract_items(chunk):
                rows_l.append({**meta, "Lesson Learned": item})

        for chunk in rep.get("recommendations", []):
            for item in extract_items(chunk):
                rows_r.append({**meta, "Recommendation": item})

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_l = pd.DataFrame(rows_l) if rows_l else pd.DataFrame(
            columns=["Report Title", "Year", "Country", "Region", "Thematic Area", "Lesson Learned"])
        df_r = pd.DataFrame(rows_r) if rows_r else pd.DataFrame(
            columns=["Report Title", "Year", "Country", "Region", "Thematic Area", "Recommendation"])
        df_l.to_excel(writer, sheet_name="Lessons Learned",  index=False)
        df_r.to_excel(writer, sheet_name="Recommendations",  index=False)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# Login page
# ─────────────────────────────────────────────────────────────────────────────

def show_login_page():
    st.markdown('<div class="login-wrap">', unsafe_allow_html=True)
    st.markdown("""
    <div class="login-logo">
      <div class="big">UNIDO</div>
      <div class="sub">Evaluation Intelligence Platform<br>IEU / EIO Division · Internal Access</div>
    </div>
    """, unsafe_allow_html=True)

    _, col, _ = st.columns([1, 2, 1])
    with col:
        st.markdown("#### Sign in")
        username = st.text_input("Username", placeholder="your.username", key="li_u")
        password = st.text_input("Password", type="password", placeholder="••••••••", key="li_p")
        if st.button("Sign in", type="primary", use_container_width=True):
            if not username or not password:
                st.error("Enter username and password.")
            else:
                with st.spinner("Authenticating…"):
                    ok, msg = do_login(username, password)
                if ok:
                    st.rerun()
                else:
                    st.error(f" {msg}")
        st.divider()
        st.caption("Access issues? Contact Maryam Babar (EIO Division).")
    st.markdown("</div>", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# SDG legend (all 17 icons displayed horizontally)
# ─────────────────────────────────────────────────────────────────────────────

def show_sdg_legend():
    badges = "".join(
        f'<span title="SDG {n}: {SDG_NAMES[n]}">'
        f'{sdg_badge_html(n, 38)}'
        f'</span>'
        for n in range(1, 18)
    )
    st.markdown(
        f'<div style="display:flex;flex-wrap:wrap;gap:2px;margin-bottom:0.5rem;">'
        f'{badges}</div>',
        unsafe_allow_html=True,
    )

# ─────────────────────────────────────────────────────────────────────────────
# Report Detail Modal
# ─────────────────────────────────────────────────────────────────────────────

def _load_ai_extraction(report_id: str) -> dict:
    """Load Gemini-extracted JSON for a report, or {} if not yet generated."""
    import pathlib
    p = pathlib.Path(__file__).parent.parent / "data" / "ai_extractions" / f"{report_id}.json"
    if p.exists():
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


@st.dialog("Evaluation Report", width="large")
def _report_detail_modal():
    """Professional modal matching the EIO platform design."""
    rep      = st.session_state.get("modal_rep", {})
    sec_data = st.session_state.get("modal_sec", {})

    title    = rep.get("title", "Untitled Report")
    year     = rep.get("year", "")
    country  = rep.get("country", "")
    region   = rep.get("region", "")
    thematic = rep.get("thematic_category", "")
    rating   = rep.get("evaluation_rating")
    donor    = rep.get("donor", "")
    budget   = rep.get("budget_usd")
    rtype    = rep.get("report_type", "Terminal Evaluation")
    sdgs     = rep.get("sdgs") or []
    rid      = rep.get("report_id", "")

    # Load Gemini AI extraction (if script has been run)
    ai = _load_ai_extraction(rid)

    # Override thematic area from AI extraction if available
    if ai.get("primary_thematic_area"):
        thematic = ai["primary_thematic_area"]
    if ai.get("context", {}).get("evaluation_rating"):
        rating = ai["context"]["evaluation_rating"]

    # Format fields
    project_id = ai.get("context", {}).get("project_id") or rep.get("project_id", "")
    rating_str = f" {float(rating):.1f} / 6" if rating else "N/A"
    budget_str = f"USD {budget/1e6:.1f}M" if budget and budget >= 1e5 else ""
    meta_line = " &nbsp;·&nbsp; ".join(p for p in [
        f" {country}" if country else "",
        f" {float(rating):.1f}/6" if rating else "",
        budget_str,
        donor,
    ] if p)
    project_line = f"Project ID: {project_id}" if project_id else ""

    # ── Dark-blue header ──────────────────────────────────────────────────────
    st.markdown(f"""
    <div style="background:#003da5;color:white;
                margin:-1rem -1rem 0 -1rem;padding:1.2rem 1.6rem 1rem;">
      <div style="font-size:0.72rem;opacity:0.75;letter-spacing:0.06em;
                  text-transform:uppercase;margin-bottom:0.4rem;">
        {rtype} &nbsp;·&nbsp; {year} &nbsp;·&nbsp; {region}{f" &nbsp;·&nbsp; {project_line}" if project_line else ""}
      </div>
      <div style="font-size:1.15rem;font-weight:700;line-height:1.35;">{title}</div>
      <div style="margin-top:0.55rem;font-size:0.82rem;opacity:0.88;">{meta_line}</div>
    </div>
    """, unsafe_allow_html=True)

    # ── Tabs ──────────────────────────────────────────────────────────────────
    # Load sections from local file first (reliable), fallback to API data
    _local_data = _load_sections_local(rid)
    if _local_data.get("sections"):
        sections = _local_data["sections"]
    else:
        sections = sec_data.get("sections", {}) if sec_data else {}
    colors   = _section_colors()

    t_ov, t_find, t_conc, t_ll, t_rec, t_sdg, t_theme, t_ctx = st.tabs(
        ["Overview", "Findings", "Conclusions", "Lessons Learned", "Recommendations", "SDG Mapping", "Thematic Area", "Context"]
    )

    with t_ov:
        st.markdown("#### AI Executive Summary")
        summary = ai.get("executive_summary") or rep.get("executive_summary") or sections.get("findings", "")
        if summary:
            ai_badge = '<span style="background:#e0f2fe;color:#0369a1;font-size:0.65rem;font-weight:700;padding:2px 7px;border-radius:4px;letter-spacing:0.06em;margin-right:8px;">✨ AI GENERATED</span>' if ai.get("executive_summary") else ""
            st.markdown(
                f'<div style="background:#f8faff;border-left:4px solid #003da5;'
                f'border-radius:0 8px 8px 0;padding:0.9rem 1.1rem;'
                f'font-size:0.88rem;line-height:1.7;color:#1e293b;">'
                f'{ai_badge}{summary[:2000]}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.info("Executive summary not yet available. Run `python scripts/extract_sections_local.py` to generate content.", icon="ℹ️")

        st.markdown("#### Classification")
        if sdgs:
            st.markdown(sdg_badges_row(sdgs, 38), unsafe_allow_html=True)
        tag_line = ""
        if thematic:
            tag_line += f'<span class="tag tag-blue">{thematic}</span> '
        if rating:
            try:
                r = float(rating)
                cls = "tag-green" if r >= 4.5 else ("tag-amber" if r >= 3.0 else "tag-red")
                label = "Satisfactory" if r >= 4.0 else ("Moderately Satisfactory" if r >= 3.0 else "Unsatisfactory")
                tag_line += f'<span class="tag {cls}">{label}</span> '
            except Exception:
                pass
        if tag_line:
            st.markdown(tag_line, unsafe_allow_html=True)

        if ai.get("thematic_justification"):
            st.markdown(
                f'<div style="margin-top:0.7rem;font-size:0.78rem;color:#6b7280;'
                f'font-style:italic;padding-left:0.5rem;">'
                f'<strong>Why this theme:</strong> {ai["thematic_justification"]}</div>',
                unsafe_allow_html=True,
            )


    with t_find:
        find_text = sections.get("findings", "") or sections.get("results", "")
        section_label = "Key Findings" if sections.get("findings") else ("Results" if sections.get("results") else "")
        if find_text and find_text.strip():
            st.markdown(
                f'<div style="font-size:0.75rem;color:#1d4ed8;font-weight:700;'
                f'letter-spacing:0.06em;margin-bottom:0.6rem;text-transform:uppercase;">'
                f'📋 {section_label} — Extracted from Report</div>',
                unsafe_allow_html=True,
            )
            st.markdown(
                f'<div style="background:#eff6ff;border-left:4px solid #1d4ed8;'
                f'border-radius:0 8px 8px 0;padding:0.85rem 1rem;'
                f'font-size:0.86rem;line-height:1.7;color:#1e293b;white-space:pre-wrap;">'
                f'{find_text.strip()[:8000]}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.info("This report does not contain a dedicated Findings or Results section.", icon="ℹ️")

    with t_conc:
        conc_text = sections.get("conclusions", "")
        if conc_text and conc_text.strip():
            st.markdown(
                '<div style="font-size:0.75rem;color:#166534;font-weight:700;'
                'letter-spacing:0.06em;margin-bottom:0.6rem;text-transform:uppercase;">📋 Extracted from Report</div>',
                unsafe_allow_html=True,
            )
            st.markdown(
                f'<div style="background:#f0fdf4;border-left:4px solid #166534;'
                f'border-radius:0 8px 8px 0;padding:0.85rem 1rem;'
                f'font-size:0.86rem;line-height:1.7;color:#1e293b;white-space:pre-wrap;">'
                f'{conc_text.strip()[:8000]}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.info("Conclusions not yet extracted for this report.", icon="ℹ️")

    with t_ll:
        # Priority: JSON file extraction → PILOT_METADATA fallback → Qdrant sections
        ai_lessons = ai.get("lessons_learned", []) or PILOT_METADATA.get(rid, {}).get("lessons_learned", [])
        if ai_lessons:
            for i, lesson in enumerate(ai_lessons, 1):
                st.markdown(
                    f'<div style="background:#e0f2fe;border-left:3px solid #0369a1;'
                    f'border-radius:0 8px 8px 0;padding:0.7rem 1rem;margin-bottom:0.5rem;'
                    f'font-size:0.86rem;line-height:1.6;color:#0c4a6e;">'
                    f'<strong>{i}.</strong> {lesson}</div>',
                    unsafe_allow_html=True,
                )
        else:
            c, bg = colors["lessons_learned"]
            _render_section_block("Lessons Learned", sections.get("lessons_learned", ""), c, bg)

    with t_rec:
        # Priority: JSON file extraction → PILOT_METADATA fallback → Qdrant sections
        ai_recs = ai.get("recommendations", []) or PILOT_METADATA.get(rid, {}).get("recommendations", [])
        if ai_recs:
            for i, rec in enumerate(ai_recs, 1):
                st.markdown(
                    f'<div style="background:#fff7ed;border-left:3px solid #ea580c;'
                    f'border-radius:0 8px 8px 0;padding:0.7rem 1rem;margin-bottom:0.5rem;'
                    f'font-size:0.86rem;line-height:1.6;color:#7c2d12;">'
                    f'<strong>{i}.</strong> {rec}</div>',
                    unsafe_allow_html=True,
                )
        else:
            c, bg = colors["recommendations"]
            _render_section_block("Recommendations", sections.get("recommendations", ""), c, bg)

    with t_sdg:
        # Use AI extraction first, then fall back to PILOT_METADATA justifications
        ai_sdg_map = ai.get("sdg_mapping", {})
        pilot_sdg_just = PILOT_METADATA.get(rid, {}).get("sdg_justifications", {})
        # Merge: prefer AI extraction; use pilot fallback if not present
        combined_sdg_map = {**pilot_sdg_just, **{str(k): v for k, v in ai_sdg_map.items()}}
        all_sdg_nums = sorted(set(
            [int(k) for k in combined_sdg_map.keys() if str(k).isdigit()] +
            ([int(s) for s in sdgs if str(s).isdigit()] if sdgs else [])
        ))
        if all_sdg_nums:
            st.markdown(
                '<div style="font-size:0.7rem;color:#166534;font-weight:700;'
                'letter-spacing:0.06em;margin-bottom:0.8rem;">✨ AI MAPPED WITH JUSTIFICATIONS</div>',
                unsafe_allow_html=True,
            )
            st.caption(
                "SDGs are inferred from project activities and outcomes — they are never explicitly "
                "stated in the report text. Each justification explains the evidence basis."
            )
            for n in all_sdg_nums:
                badge = sdg_badge_html(n, 42)
                justification = combined_sdg_map.get(str(n), "")
                just_html = (
                    f'<div style="font-size:0.80rem;color:#374151;margin-top:4px;line-height:1.55;">'
                    f'{justification}</div>'
                ) if justification else ""
                st.markdown(
                    f'<div style="display:flex;align-items:flex-start;gap:14px;'
                    f'background:#f0fdf4;border:1px solid #bbf7d0;border-radius:8px;'
                    f'padding:0.75rem 1rem;margin-bottom:0.5rem;">'
                    f'{badge}'
                    f'<div><strong style="font-size:0.9rem;color:#166534;">'
                    f'SDG {n} — {SDG_NAMES.get(n,"")}</strong>'
                    f'{just_html}</div></div>',
                    unsafe_allow_html=True,
                )
        else:
            st.info("No SDG mapping available for this report.")

    with t_theme:
        # Use AI extraction first, then PILOT_METADATA fallback
        pilot_meta   = PILOT_METADATA.get(rid, {})
        theme_name   = ai.get("primary_thematic_area") or pilot_meta.get("thematic_category") or thematic
        theme_second = ai.get("secondary_thematic_area") or pilot_meta.get("secondary_thematic_area", "")
        theme_just   = ai.get("thematic_justification") or pilot_meta.get("thematic_justification", "")

        if theme_name:
            st.markdown(
                '<div style="font-size:0.7rem;color:#7c3aed;font-weight:700;'
                'letter-spacing:0.06em;margin-bottom:0.8rem;">✨ AI CLASSIFIED WITH JUSTIFICATION</div>',
                unsafe_allow_html=True,
            )
            st.caption(
                "Thematic areas are inferred from project objectives, activities, and sector focus — "
                "they are not stated explicitly in the report. Justification explains the evidence basis."
            )
            # Primary theme card
            st.markdown(
                f'<div style="background:#f5f3ff;border:1px solid #ddd6fe;border-radius:10px;'
                f'padding:1rem 1.2rem;margin-bottom:0.75rem;">'
                f'<div style="font-size:0.68rem;font-weight:700;color:#7c3aed;'
                f'letter-spacing:0.07em;text-transform:uppercase;margin-bottom:0.3rem;">'
                f'Primary Thematic Area</div>'
                f'<div style="font-size:1.05rem;font-weight:700;color:#4c1d95;margin-bottom:0.5rem;">'
                f'{theme_name}</div>'
                + (
                    f'<div style="font-size:0.83rem;color:#374151;line-height:1.6;">'
                    f'<strong>Why this theme:</strong> {theme_just}</div>'
                    if theme_just else ""
                )
                + '</div>',
                unsafe_allow_html=True,
            )
            # Secondary theme (if any)
            if theme_second:
                st.markdown(
                    f'<div style="background:#faf5ff;border:1px solid #e9d5ff;border-radius:8px;'
                    f'padding:0.75rem 1rem;">'
                    f'<div style="font-size:0.68rem;font-weight:700;color:#9333ea;'
                    f'letter-spacing:0.07em;text-transform:uppercase;margin-bottom:0.3rem;">'
                    f'Secondary Thematic Area</div>'
                    f'<div style="font-size:0.95rem;font-weight:600;color:#6b21a8;">'
                    f'{theme_second}</div></div>',
                    unsafe_allow_html=True,
                )
            # Show all UNIDO thematic areas with current one highlighted
            st.markdown(
                "<div style='margin-top:1rem;font-size:0.78rem;font-weight:600;"
                "color:#6b7280;margin-bottom:0.4rem;'>UNIDO Thematic Framework</div>",
                unsafe_allow_html=True,
            )
            chips_html = ""
            for area in THEMATIC_AREAS:
                is_primary   = area == theme_name
                is_secondary = area == theme_second
                if is_primary:
                    style = ("background:#7c3aed;color:white;font-weight:700;"
                             "border-radius:999px;padding:0.2rem 0.7rem;font-size:0.76rem;"
                             "display:inline-block;margin:3px;")
                elif is_secondary:
                    style = ("background:#ddd6fe;color:#4c1d95;font-weight:600;"
                             "border-radius:999px;padding:0.2rem 0.7rem;font-size:0.76rem;"
                             "display:inline-block;margin:3px;")
                else:
                    style = ("background:#f3f4f6;color:#9ca3af;"
                             "border-radius:999px;padding:0.2rem 0.7rem;font-size:0.76rem;"
                             "display:inline-block;margin:3px;")
                chips_html += f'<span style="{style}">{area}</span>'
            st.markdown(f'<div style="line-height:2;">{chips_html}</div>', unsafe_allow_html=True)
        else:
            st.info("Thematic area classification not yet available for this report.")

    with t_ctx:
        # Funding line
        donor_val = rep.get("donor", "")
        funding_text = f"Funded by <strong>{donor_val}</strong>. Implemented in {country}." if donor_val and country else (f"Implemented in {country}." if country else "")
        if funding_text:
            st.markdown(
                f'<div style="font-size:0.85rem;color:#374151;margin-bottom:1rem;'
                f'padding-bottom:0.8rem;border-bottom:1px solid #f0f0f0;">{funding_text}</div>',
                unsafe_allow_html=True,
            )

        # Rating label from metadata
        rating_lbl = rep.get("overall_rating_label") or "N/A"
        if not rating_lbl and rating:
            try:
                r_val = float(rating)
                if r_val >= 4.5:    rating_lbl = "Highly Satisfactory"
                elif r_val >= 4.0:  rating_lbl = "Satisfactory"
                elif r_val >= 3.0:  rating_lbl = "Moderately Satisfactory"
                else:               rating_lbl = "Unsatisfactory"
            except Exception:
                pass

        def _info_card(label, value):
            return (
                f'<div style="background:white;border:1px solid #e5e7eb;border-radius:8px;padding:0.75rem 1rem;">' 
                f'<div style="font-size:0.64rem;font-weight:700;color:#9ca3af;'
                f'letter-spacing:0.08em;text-transform:uppercase;margin-bottom:0.35rem;">{label}</div>'
                f'<div style="font-size:0.9rem;font-weight:700;color:#111827;line-height:1.3;">{value}</div>'
                f'</div>'
            )

        # Row 1: Budget | Year | Thematic Area
        r1c1, r1c2, r1c3 = st.columns(3)
        with r1c1:
            st.markdown(_info_card("Budget", budget_str or "N/A"), unsafe_allow_html=True)
        with r1c2:
            st.markdown(_info_card("Year", str(year) if year else "N/A"), unsafe_allow_html=True)
        with r1c3:
            st.markdown(_info_card("Thematic Area", rep.get("thematic_category","N/A")), unsafe_allow_html=True)

        st.markdown("<div style='height:8px;'></div>", unsafe_allow_html=True)

        # Row 2: Type | Country | Overall Rating
        r2c1, r2c2, r2c3 = st.columns(3)
        with r2c1:
            st.markdown(_info_card("Evaluation Type", rtype or "N/A"), unsafe_allow_html=True)
        with r2c2:
            st.markdown(_info_card("Country", country or "N/A"), unsafe_allow_html=True)
        with r2c3:
            st.markdown(_info_card("Overall Rating", rating_lbl), unsafe_allow_html=True)

    # ── Footer downloads ──────────────────────────────────────────────────────
    st.divider()
    if sec_data and sec_data.get("sections"):
        xl = make_excel_sections([rep], {rid: sec_data})
        st.download_button(
            "⬇ Download Excel",
            data=xl,
            file_name=f"UNIDO_{rid}_Evaluation.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )


# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — Search & Browse
# ─────────────────────────────────────────────────────────────────────────────

def _rating_badge(rating) -> str:
    if rating is None:
        return '<span class="rating-pill rating-none">Rating: N/A</span>'
    try:
        r = float(rating)
    except (TypeError, ValueError):
        return '<span class="rating-pill rating-none">Rating: N/A</span>'
    cls = "rating-high" if r >= 4.5 else ("rating-mid" if r >= 3.0 else "rating-low")
    return f'<span class="rating-pill {cls}"> {r:.1f} / 6</span>'


def _section_colors() -> dict:
    return {
        "findings":        ("#1d4ed8", "#eff6ff"),
        "results":         ("#7c3aed", "#f5f3ff"),
        "lessons_learned": ("#0369a1", "#e0f2fe"),
        "conclusions":     ("#166534", "#f0fdf4"),
        "recommendations": ("#9a3412", "#fff7ed"),
    }


def _render_section_block(label: str, text: str, color: str, bg: str):
    if not text or not text.strip():
        st.markdown(
            f'<div style="color:#9ca3af;font-size:0.82rem;font-style:italic;padding:0.4rem 0;">' +
            f'No {label.lower()} content extracted from this report.</div>',
            unsafe_allow_html=True,
        )
        return
    display_text = text.strip()
    if len(display_text) > 8000:
        display_text = display_text[:8000] + "\n\n[… truncated — download Excel for full text]"
    st.markdown(
        f'<div style="background:{bg};border-left:4px solid {color};' +
        f'border-radius:0 8px 8px 0;padding:0.85rem 1rem;margin-top:0.3rem;">' +
        f'<div style="font-size:0.84rem;color:#1e293b;line-height:1.7;white-space:pre-wrap;">' +
        display_text + '</div></div>',
        unsafe_allow_html=True,
    )


def show_search_tab(filters: dict):
    load_reports()
    # Demo phase: restrict to the 4 verified pilot reports only
    all_reps = [r for r in (st.session_state.all_reports or [])
                if r.get("report_id") in PILOT_METADATA]
    all_reps = sorted(all_reps, key=lambda r: (r.get("year") or 0, r.get("report_id", "")), reverse=True)

    st.markdown(
        '<div class="pilot-banner"><strong>Pilot phase</strong> — '
        'Showing 4 evaluation reports from 2021. Sections verified and ready for review.</div>',
        unsafe_allow_html=True,
    )

    # ── Apply sidebar filters ────────────────────────────────────────────────
    filtered = all_reps
    if filters["thematic"]:
        filtered = [r for r in filtered if r.get("thematic_category") in filters["thematic"]]
    if filters["sdgs"]:
        filtered = [r for r in filtered
                    if any(s in (r.get("sdgs") or []) for s in filters["sdgs"])]
    if filters["eval_type"]:
        filtered = [r for r in filtered if (r.get("report_type") or r.get("evaluation_type") or "") in filters["eval_type"]]
    if filters["region"]:
        filtered = [r for r in filtered if r.get("region") in filters["region"]]
    if filters.get("year_min"):
        filtered = [r for r in filtered if r.get("year") == filters["year_min"]]

    if len(all_reps) == 0:
        st.info("No reports loaded. Check backend connection.")
        return
    st.markdown(
        f'<div style="font-size:0.82rem;color:#6b7280;margin-bottom:0.8rem;">'
        f'Showing <strong>{len(filtered)}</strong> of <strong>{len(all_reps)}</strong> reports</div>',
        unsafe_allow_html=True,
    )
    if not filtered:
        return

    # ── 2-column card grid ───────────────────────────────────────────────────
    for row_start in range(0, len(filtered), 2):
        pair = filtered[row_start : row_start + 2]
        grid_cols = st.columns(len(pair), gap="large")

        for col, rep in zip(grid_cols, pair):
            title    = rep.get("title") or "Untitled Report"
            year     = rep.get("year", "")
            country  = rep.get("country", "")
            region   = rep.get("region", "")
            thematic = rep.get("thematic_category", "")
            rtype    = rep.get("report_type", "")
            donor    = rep.get("donor", "")
            sdgs     = rep.get("sdgs") or []
            rid      = rep.get("report_id", "")
            rating   = rep.get("evaluation_rating")
            budget   = rep.get("budget_usd")
            project_id = rep.get("project_id", "")

            badges_html    = sdg_badges_row(sdgs, 28)
            rating_html    = _rating_badge(rating)
            thematic_badge = f'<span class="tag tag-blue">{thematic}</span>' if thematic else ""
            rtype_badge    = f'<span class="tag tag-gray">{rtype}</span>' if rtype else ""
            donor_badge    = f'<span class="tag tag-green">{donor}</span>' if donor else ""
            budget_str     = (f'<span class="tag tag-purple">USD {budget/1e6:.1f}M</span>'
                             if budget and budget >= 1e5 else "")

            meta_parts = [str(year)] if year else []
            if country: meta_parts.append(country)
            if region and region != country: meta_parts.append(region)
            if project_id: meta_parts.append(f"Project {project_id}")
            meta_line = " · ".join(meta_parts)

            with col:
                st.markdown(f"""
                <div class="report-card" style="height:100%;min-height:160px;">
                  <div style="font-size:0.68rem;font-weight:700;color:#009EDB;
                              letter-spacing:0.06em;text-transform:uppercase;
                              margin-bottom:0.4rem;">{meta_line}</div>
                  <div class="report-title" style="font-size:0.9rem;line-height:1.45;
                              margin-bottom:0.5rem;">{title}</div>
                  <div style="margin:0.3rem 0 0.4rem;display:flex;flex-wrap:wrap;gap:3px;">
                    {rating_html}{thematic_badge}{rtype_badge}{donor_badge}{budget_str}
                  </div>
                  <div style="margin-top:0.35rem;">{badges_html}</div>
                </div>
                """, unsafe_allow_html=True)

                # ── Action buttons — 3 equal columns, full width of card ───
                b_view, b_ai, b_exp = st.columns(3, gap="small")
                with b_view:
                    if st.button("View Details ↗", key=f"view_{rid}", use_container_width=True):
                        with st.spinner("Loading…"):
                            sec_data = load_sections(rid)
                        st.session_state["modal_rep"] = rep
                        st.session_state["modal_sec"] = sec_data
                        _report_detail_modal()
                with b_ai:
                    askai_key = f"askai_prompt_{rid}"
                    if st.session_state.get(askai_key):
                        # Show the pre-built prompt panel
                        st.markdown(
                            f'<a href="https://copilot.microsoft.com" target="_blank" '
                            f'style="display:block;text-align:center;background:#0078d4;color:white;'
                            f'padding:6px 10px;border-radius:6px;font-size:0.82rem;font-weight:600;'
                            f'text-decoration:none;margin-bottom:4px;">🤖 Open Copilot ↗</a>',
                            unsafe_allow_html=True,
                        )
                        st.text_area(
                            "Copy → paste into Copilot:",
                            value=st.session_state[askai_key],
                            height=160,
                            key=f"askai_ta_{rid}",
                            label_visibility="visible",
                        )
                        if st.button("✕ Close", key=f"askai_close_{rid}", use_container_width=True):
                            del st.session_state[askai_key]
                            st.rerun()
                    else:
                        if st.button("Ask AI", key=f"askai_{rid}", type="primary", use_container_width=True):
                            ai_data = _load_ai_extraction(rid)
                            ctx     = ai_data.get("context", {})
                            _title  = ctx.get("title") or rep.get("title", "")
                            _year   = ctx.get("year") or rep.get("year", "")
                            _cntry  = ctx.get("country") or rep.get("country", "")
                            _theme  = ai_data.get("primary_thematic_area") or rep.get("thematic_category", "")
                            _summ   = ai_data.get("executive_summary") or rep.get("executive_summary", "")
                            _les    = ai_data.get("lessons_learned") or []
                            _recs   = ai_data.get("recommendations") or []
                            _sdgs   = ai_data.get("sdg_mapping") or {}
                            _sdg_s  = ", ".join(
                                f"SDG {k}" for k in sorted(_sdgs.keys(), key=lambda x: int(x) if x.isdigit() else 99)
                            ) if _sdgs else ""
                            parts = [
                                f"UNIDO Evaluation Report: {_title} ({_year}, {_cntry})",
                                f"Thematic Area: {_theme}" if _theme else "",
                                f"SDGs: {_sdg_s}" if _sdg_s else "",
                                "",
                                "EXECUTIVE SUMMARY:",
                                (_summ[:1500] if _summ else "(not available)"),
                            ]
                            if _les:
                                parts += ["", "KEY LESSONS LEARNED:"]
                                parts += [f"{i}. {l}" for i, l in enumerate(_les[:5], 1)]
                            if _recs:
                                parts += ["", "KEY RECOMMENDATIONS:"]
                                parts += [f"{i}. {r}" for i, r in enumerate(_recs[:5], 1)]
                            parts += ["", "---", "Please analyse this evaluation report and share your insights."]
                            st.session_state[askai_key] = "\n".join(p for p in parts)
                            st.rerun()
                with b_exp:
                    exp_key = f"export_bytes_{rid}"
                    exp_err_key = f"export_err_{rid}"
                    if exp_key in st.session_state:
                        st.download_button(
                            label="⬇ Excel",
                            data=st.session_state[exp_key],
                            file_name=f"UNIDO_{rid}_Evaluation.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_{rid}",
                            use_container_width=True,
                        )
                    else:
                        if st.session_state.get(exp_err_key):
                            st.error(st.session_state.pop(exp_err_key), icon="⚠️")
                        if st.button("Export", key=f"exp_{rid}", use_container_width=True):
                            with st.spinner("Preparing Excel…"):
                                try:
                                    sec_e = load_sections(rid)
                                    xl_bytes = make_excel_sections([rep], {rid: sec_e})
                                    st.session_state[exp_key] = xl_bytes
                                except Exception as _ex:
                                    st.session_state[exp_err_key] = f"Export failed: {_ex}"
                            st.rerun()

        st.markdown("<div style='height:0.8rem;'></div>", unsafe_allow_html=True)



# ─────────────────────────────────────────────────────────────────────────────
# Synthesis History — persistent disk storage
# ─────────────────────────────────────────────────────────────────────────────

_SYNTH_HISTORY_PATH = pathlib.Path(__file__).parent.parent / "data" / "synthesis_history.json"

def _load_synth_history() -> list:
    """Load all saved Q&A pairs from disk. Returns list, newest-first."""
    try:
        if _SYNTH_HISTORY_PATH.exists():
            with open(_SYNTH_HISTORY_PATH, "r", encoding="utf-8") as _f:
                data = json.load(_f)
            return data if isinstance(data, list) else []
    except Exception:
        pass
    return []

def _save_synth_history(history: list):
    """Persist Q&A history list to disk."""
    try:
        _SYNTH_HISTORY_PATH.parent.mkdir(parents=True, exist_ok=True)
        with open(_SYNTH_HISTORY_PATH, "w", encoding="utf-8") as _f:
            json.dump(history, _f, ensure_ascii=False, indent=2)
    except Exception as _e:
        st.warning(f"Could not save history: {_e}")

def _delete_synth_item(idx: int):
    """Remove one Q&A entry by index and persist."""
    h = _load_synth_history()
    if 0 <= idx < len(h):
        h.pop(idx)
        _save_synth_history(h)

# TAB 2 — Synthesis (RAG — passages only, no LLM)
# ─────────────────────────────────────────────────────────────────────────────

def show_synthesis_tab(filters: dict):
    load_reports()
    # Demo phase: restrict to the 4 verified pilot reports only
    all_reps = [r for r in (st.session_state.all_reports or [])
                if r.get("report_id") in PILOT_METADATA]

    # ── Sub-tabs: Ask Claude | History ───────────────────────────────────────
    sub_ask, sub_hist = st.tabs(["Ask AI", "History"])

    # ════════════════════════════════════════════════════════════════════════
    # SUB-TAB 1 — Ask Claude
    # ════════════════════════════════════════════════════════════════════════
    with sub_ask:
        col_sel, col_chat = st.columns([3, 7])

        # ── Report selector ──────────────────────────────────────────────────
        with col_sel:
            st.markdown("#### Select Reports")
            search_term = st.text_input("Filter list", placeholder="Search title…", key="synth_search")
            visible = [r for r in all_reps
                       if not search_term or search_term.lower() in r.get("title","").lower()]

            c1, c2 = st.columns(2)
            with c1:
                if st.button("Select all", use_container_width=True):
                    st.session_state["synth_sel"] = [r["report_id"] for r in visible]
            with c2:
                if st.button("Clear all", use_container_width=True):
                    st.session_state["synth_sel"] = []

            if "synth_sel" not in st.session_state:
                st.session_state["synth_sel"] = []

            selected_ids = st.session_state["synth_sel"]
            st.caption(f"{len(selected_ids)} selected")

            for rep in visible:
                rid = rep["report_id"]
                checked = rid in selected_ids
                label = f"{rep.get('year','')} — {display_title(rep)}"
                new = st.checkbox(label, value=checked, key=f"synth_cb_{rid}")
                if new and rid not in selected_ids:
                    selected_ids.append(rid)
                elif not new and rid in selected_ids:
                    selected_ids.remove(rid)

        # ── Chat area ────────────────────────────────────────────────────────
        with col_chat:
            st.markdown("#### Ask a Question Across Reports")

            if not selected_ids:
                st.info("Select one or more reports on the left to begin.")
            else:
                # Selected report pills
                sel_reps = [r for r in all_reps if r["report_id"] in selected_ids]
                pills_html = " ".join(
                    f'<span class="tag tag-blue">{display_title(r)}</span>'
                    for r in sel_reps[:8]
                )
                if len(sel_reps) > 8:
                    pills_html += f' <span class="tag tag-gray">+{len(sel_reps)-8} more</span>'
                st.markdown(pills_html, unsafe_allow_html=True)
                st.divider()

                # Example chips
                examples = [
                    "What are the key lessons learned across these reports?",
                    "What common recommendations emerge on project sustainability?",
                    "What findings relate to clean energy effectiveness?",
                    "How do these projects address gender mainstreaming?",
                    "What conclusions were drawn about stakeholder engagement?",
                    "Compare the key findings across all selected reports.",
                ]
                st.markdown("**Example queries:**")
                ex_cols = st.columns(2)
                for i, ex in enumerate(examples):
                    with ex_cols[i % 2]:
                        if st.button(ex, key=f"synth_ex_{i}"):
                            st.session_state["synth_prefill"] = ex

                query = st.text_area(
                    "Question",
                    value=st.session_state.pop("synth_prefill", ""),
                    height=90,
                    placeholder="What findings emerge across the selected reports on…",
                    label_visibility="collapsed",
                    key="synth_query",
                )

                send = st.button("Ask AI", type="primary", use_container_width=False)

                if send and query.strip():
                    with st.spinner("Claude is synthesising across selected reports…"):
                        try:
                            import anthropic as _anthropic

                            try:
                                _api_key = st.secrets["ANTHROPIC_API_KEY"]
                            except Exception:
                                _api_key = os.getenv("ANTHROPIC_API_KEY", "")

                            if not _api_key:
                                st.warning("Add ANTHROPIC_API_KEY to Streamlit secrets to enable synthesis.", icon="⚠️")
                                st.stop()

                            context_blocks = []
                            for rid in selected_ids:
                                sec   = _load_sections_local(rid)
                                secs  = sec.get("sections", {})
                                meta  = sec.get("metadata", {})
                                ai_d  = _load_ai_extraction(rid)
                                title   = meta.get("title") or ai_d.get("context", {}).get("title", rid)
                                year    = meta.get("year") or ai_d.get("context", {}).get("year", "")
                                country = meta.get("country") or ai_d.get("context", {}).get("country", "")
                                block = (
                                    f"=== REPORT: {title} ===\n"
                                    f"Year: {year} | Country: {country}\n\n"
                                    f"FINDINGS / RESULTS:\n{(secs.get('findings') or secs.get('results') or 'Not available')[:1500]}\n\n"
                                    f"CONCLUSIONS:\n{secs.get('conclusions', 'Not available')[:1500]}\n\n"
                                    f"LESSONS LEARNED:\n{secs.get('lessons_learned', 'Not available')[:1500]}\n\n"
                                    f"RECOMMENDATIONS:\n{secs.get('recommendations', 'Not available')[:1500]}\n"
                                )
                                context_blocks.append(block)

                            n = len(context_blocks)
                            context_text = "\n\n".join(context_blocks)

                            system_prompt = (
                                f"You are a senior evaluation synthesis analyst at UNIDO's Independent Evaluation Unit (IEU). "
                                f"You have deep expertise in development effectiveness, OECD-DAC criteria, clean energy, and climate action evaluation.\n\n"
                                f"You have been provided with {n} UNIDO evaluation report(s). "
                                f"Synthesise findings ACROSS all reports — analytical, evidence-based, and genuinely useful.\n\n"
                                f"RULES:\n"
                                f"1. Synthesise ACROSS reports — identify patterns, tensions, and cross-cutting themes\n"
                                f"2. Use precise language: 'in 3 of 4 reports', 'the majority of reports'\n"
                                f"3. Cite specific report titles when making factual claims\n"
                                f"4. Draw non-obvious analytical conclusions\n"
                                f"5. Never invent information not in the provided reports\n"
                                f"6. End EVERY response with a ## Key Findings section with 3–5 bullet points\n\n"
                                f"TONE: Senior UN evaluation expert writing for a peer audience. Direct, analytical, clear."
                            )

                            _client = _anthropic.Anthropic(api_key=_api_key)
                            msg = _client.messages.create(
                                model="claude-sonnet-4-5",
                                max_tokens=2048,
                                system=system_prompt,
                                messages=[{"role": "user", "content": f"EVALUATION REPORTS:\n\n{context_text}\n\nQUESTION: {query.strip()}"}]
                            )
                            answer = msg.content[0].text if msg.content else ""

                            import datetime as _dt
                            new_entry = {
                                "query": query.strip(),
                                "answer": answer,
                                "report_count": n,
                                "report_ids": list(selected_ids),
                                "report_titles": [r.get("title","") for r in sel_reps],
                                "timestamp": _dt.datetime.now().strftime("%Y-%m-%d %H:%M"),
                            }
                            # Save to disk (prepend so newest is first)
                            history = _load_synth_history()
                            history.insert(0, new_entry)
                            _save_synth_history(history)

                            # Show answer immediately in this tab
                            st.markdown(
                                f'<div style="background:#f0f4ff;border-left:3px solid #003da5;'
                                f'padding:0.5rem 0.8rem;border-radius:0 6px 6px 0;'
                                f'font-size:0.85rem;font-weight:600;color:#003da5;margin-bottom:0.5rem;">'
                                f'❓ {new_entry["query"]}</div>',
                                unsafe_allow_html=True,
                            )
                            st.markdown(answer)
                            st.caption(f"Synthesised across {n} report(s) · Powered by Claude · Saved to History")

                        except Exception as e:
                            st.error(f"Synthesis error: {e}")

    # ════════════════════════════════════════════════════════════════════════
    # SUB-TAB 2 — History
    # ════════════════════════════════════════════════════════════════════════
    with sub_hist:
        st.markdown("#### Synthesis History")
        st.caption("All questions ever asked — reload anytime without using AI credits.")

        history = _load_synth_history()

        if not history:
            st.info("No questions saved yet. Ask a question in the **Ask AI** tab and it will appear here.")
        else:
            # Search / filter
            hist_search = st.text_input("🔍 Search questions", placeholder="Filter by keyword…", key="hist_search")
            if hist_search:
                history = [h for h in history if hist_search.lower() in h.get("query","").lower()
                           or hist_search.lower() in h.get("answer","").lower()]
                st.caption(f"{len(history)} result(s) matching '{hist_search}'")

            st.markdown(f"**{len(history)} saved question(s)**")
            st.divider()

            for i, item in enumerate(history):
                ts   = item.get("timestamp", "")
                q    = item.get("query", "")
                ans  = item.get("answer", "")
                n_r  = item.get("report_count", 0)
                rpts = item.get("report_titles", [])

                col_q, col_del = st.columns([10, 1])
                with col_q:
                    st.markdown(
                        f'<div style="background:#f0f4ff;border-left:3px solid #003da5;'
                        f'padding:0.5rem 0.8rem;border-radius:0 6px 6px 0;'
                        f'font-size:0.9rem;font-weight:600;color:#003da5;">'
                        f'❓ {q}</div>',
                        unsafe_allow_html=True,
                    )
                with col_del:
                    if st.button("🗑️", key=f"del_hist_{i}", help="Delete this entry"):
                        _delete_synth_item(i)
                        st.rerun()

                # Metadata row
                meta_parts = []
                if ts:
                    meta_parts.append(f"📅 {ts}")
                if n_r:
                    meta_parts.append(f"📄 {n_r} report(s)")
                if rpts:
                    titles_short = ", ".join(t[:35] for t in rpts[:3])
                    if len(rpts) > 3:
                        titles_short += f" +{len(rpts)-3} more"
                    meta_parts.append(f"*{titles_short}*")
                st.caption(" · ".join(meta_parts))

                # Collapsible answer
                with st.expander("View answer", expanded=(i == 0)):
                    if ans:
                        st.markdown(ans)
                    else:
                        st.info("No answer recorded.")

                st.divider()

            # Bulk clear
            if st.button("🗑️ Clear all history", type="secondary"):
                _save_synth_history([])
                st.success("History cleared.")
                st.rerun()


# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — Visualize
# ─────────────────────────────────────────────────────────────────────────────

def _build_knowledge_graph_figure(reports: list, by_sdg: dict, by_thematic: dict, by_year: dict):
    """Build a Plotly knowledge graph: reports as nodes, grouped by thematic area,
    edges drawn for shared SDGs (≥2). Returns a go.Figure."""
    import math

    if not reports:
        return None

    # ── Assign theme colors ───────────────────────────────────────────────────
    theme_palette = [
        "#009EDB", "#4CAF50", "#FF9800", "#E91E63", "#9C27B0",
        "#00BCD4", "#FF5722", "#607D8B", "#795548", "#3F51B5", "#8BC34A",
    ]
    theme_list = sorted(set(r.get("thematic_category", "Unknown") for r in reports))
    theme_color = {t: theme_palette[i % len(theme_palette)] for i, t in enumerate(theme_list)}

    # ── Cluster reports by thematic area ─────────────────────────────────────
    clusters: dict = {}
    for r in reports:
        t = r.get("thematic_category", "Unknown")
        clusters.setdefault(t, []).append(r)

    n_clusters = len(clusters)
    cluster_radius = 4.5
    report_pos: dict = {}   # report_id → (x, y)
    theme_pos: dict  = {}   # theme_name → (x, y)

    for ci, (theme, reps) in enumerate(sorted(clusters.items())):
        ca = 2 * math.pi * ci / n_clusters
        cx = cluster_radius * math.cos(ca)
        cy = cluster_radius * math.sin(ca)
        theme_pos[theme] = (cx * 1.35, cy * 1.35)

        n = len(reps)
        for ri, rep in enumerate(reps):
            if n == 1:
                rx, ry = cx, cy
            else:
                ra = 2 * math.pi * ri / n
                spread = min(1.0 + n * 0.06, 1.8)
                rx = cx + spread * math.cos(ra)
                ry = cy + spread * math.sin(ra)
            report_pos[rep["report_id"]] = (rx, ry)

    # ── Theme → report edges ──────────────────────────────────────────────────
    theme_edge_x, theme_edge_y = [], []
    for theme, reps in clusters.items():
        tx, ty = theme_pos[theme]
        for rep in reps:
            rx, ry = report_pos[rep["report_id"]]
            theme_edge_x += [rx, tx, None]
            theme_edge_y += [ry, ty, None]

    # ── SDG-connection edges (reports sharing ≥2 SDGs) ───────────────────────
    sdg_edge_x, sdg_edge_y = [], []
    id_list = [r["report_id"] for r in reports]
    sdg_map = {r["report_id"]: set(r.get("sdgs") or []) for r in reports}
    for i in range(len(id_list)):
        for j in range(i + 1, len(id_list)):
            shared = sdg_map[id_list[i]] & sdg_map[id_list[j]]
            if len(shared) >= 2:
                x1, y1 = report_pos[id_list[i]]
                x2, y2 = report_pos[id_list[j]]
                sdg_edge_x += [x1, x2, None]
                sdg_edge_y += [y1, y2, None]

    # ── Build traces ──────────────────────────────────────────────────────────
    fig = go.Figure()

    # Theme-report edges (light grey)
    fig.add_trace(go.Scatter(
        x=theme_edge_x, y=theme_edge_y, mode="lines",
        line=dict(color="rgba(200,200,200,0.4)", width=1),
        hoverinfo="none", showlegend=False,
    ))

    # SDG connection edges (blue, thinner)
    if sdg_edge_x:
        fig.add_trace(go.Scatter(
            x=sdg_edge_x, y=sdg_edge_y, mode="lines",
            line=dict(color="rgba(0,158,219,0.25)", width=1.5),
            hoverinfo="none", showlegend=False, name="Shared SDGs",
        ))

    # Theme nodes (larger squares)
    for theme in sorted(clusters.keys()):
        tx, ty = theme_pos[theme]
        n_reps = len(clusters[theme])
        fig.add_trace(go.Scatter(
            x=[tx], y=[ty], mode="markers+text",
            marker=dict(
                size=22, color=theme_color[theme],
                symbol="square", line=dict(color="white", width=2),
            ),
            text=[theme.replace(" / ", "<br>").replace(" & ", "<br>")],
            textposition="top center",
            textfont=dict(size=9, color="#1a1a2e"),
            hovertext=f"<b>{theme}</b><br>{n_reps} reports",
            hoverinfo="text",
            showlegend=True,
            name=theme,
            legendgroup=theme,
        ))

    # Report nodes (colored by theme)
    rep_x = [report_pos[r["report_id"]][0] for r in reports]
    rep_y = [report_pos[r["report_id"]][1] for r in reports]
    rep_colors = [theme_color.get(r.get("thematic_category", "Unknown"), "#888") for r in reports]
    rep_hover = []
    for r in reports:
        sdgs_str = " · ".join([f"SDG {s}" for s in sorted(r.get("sdgs") or [])[:4]])
        rep_hover.append(
            f"<b>{r.get('title','')[:55]}…</b><br>"
            f"{r.get('year','')} · {r.get('country','')}<br>"
            f"{r.get('thematic_category','')}<br>"
            f"{sdgs_str}"
        )

    fig.add_trace(go.Scatter(
        x=rep_x, y=rep_y, mode="markers",
        marker=dict(
            size=10, color=rep_colors,
            line=dict(color="white", width=1.5),
            opacity=0.9,
        ),
        hovertext=rep_hover, hoverinfo="text",
        showlegend=False, name="Reports",
    ))

    fig.update_layout(
        height=620,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(249,251,252,1)",
        margin=dict(t=20, b=20, l=20, r=20),
        showlegend=True,
        legend=dict(
            orientation="v", x=1.01, y=1, bgcolor="rgba(255,255,255,0.9)",
            bordercolor="#e5e7eb", borderwidth=1,
            font=dict(size=10),
        ),
        xaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
        yaxis=dict(showgrid=False, zeroline=False, showticklabels=False),
        hovermode="closest",
    )
    return fig


def show_visualize_tab():
    load_reports()
    reports = [r for r in (st.session_state.all_reports or [])
               if r.get("report_id") in PILOT_METADATA]
    if not reports:
        st.info("No reports indexed yet.")
        return

    # ── Compute aggregates ────────────────────────────────────────────────────
    by_sdg: dict      = {}
    by_thematic: dict = {}
    for r in reports:
        t = r.get("thematic_category") or "Unknown"
        by_thematic[t] = by_thematic.get(t, 0) + 1
        for s in (r.get("sdgs") or []):
            by_sdg[f"SDG {s}"] = by_sdg.get(f"SDG {s}", 0) + 1

    sdg_count = len(by_sdg)
    countries  = list(set(r.get("country","") for r in reports if r.get("country")))
    avg_rating = sum(float(r["evaluation_rating"]) for r in reports if r.get("evaluation_rating")) / max(len([r for r in reports if r.get("evaluation_rating")]),1)

    # ── KPI Banner ────────────────────────────────────────────────────────────
    st.markdown("""
    <div style="background:#003DA5;border-radius:10px;padding:1rem 1.5rem;
                display:flex;gap:0;margin-bottom:1rem;">
    """, unsafe_allow_html=True)

    kpis = [
        (str(len(reports)), "Evaluation Reports", "#60a5fa"),
        (str(sdg_count),    "SDGs Covered",       "#34d399"),
        (str(len(countries)),"Countries",          "#fbbf24"),
        (f"{avg_rating:.1f}/6", "Avg. Rating",    "#f87171"),
        (str(len(by_thematic)),"Thematic Areas",  "#a78bfa"),
    ]
    cols = st.columns(5)
    for col, (num, lbl, color) in zip(cols, kpis):
        col.markdown(
            f'<div style="background:rgba(255,255,255,0.08);border-radius:8px;'
            f'padding:0.9rem 0.5rem;text-align:center;border-left:4px solid {color};">'
            f'<div style="font-size:1.9rem;font-weight:800;color:{color};line-height:1;">{num}</div>'
            f'<div style="font-size:0.72rem;color:#cbd5e1;margin-top:4px;text-transform:uppercase;'
            f'letter-spacing:0.05em;">{lbl}</div></div>',
            unsafe_allow_html=True,
        )

    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)

    # ── Row 1: SDG Coverage strip ─────────────────────────────────────────────
    st.markdown('<div style="font-size:0.78rem;font-weight:700;color:#003DA5;'
                'text-transform:uppercase;letter-spacing:0.06em;margin:0.5rem 0 0.4rem;">SDG Coverage</div>',
                unsafe_allow_html=True)
    sdg_strip = ""
    for n in range(1, 18):
        count = by_sdg.get(f"SDG {n}", 0)
        opacity = "1.0" if count > 0 else "0.18"
        sdg_strip += (
            f'<span style="display:inline-flex;flex-direction:column;align-items:center;'
            f'margin:2px;opacity:{opacity};" title="SDG {n}: {SDG_NAMES[n]} — {count} report(s)">'
            f'{sdg_badge_html(n, 44)}'
            f'<span style="font-size:8.5px;color:#374151;font-weight:600;margin-top:2px;">{count if count else "·"}</span>'
            f'</span>'
        )
    st.markdown(f'<div style="background:#f8faff;border-radius:8px;padding:0.6rem 0.8rem;">'
                f'<div style="display:flex;flex-wrap:wrap;gap:2px;">{sdg_strip}</div></div>',
                unsafe_allow_html=True)

    st.markdown("<div style='height:10px'></div>", unsafe_allow_html=True)

    # ── Row 2: Ratings | Thematic | SDG bar ──────────────────────────────────
    col_r, col_t, col_s = st.columns([1.1, 1, 1])

    with col_r:
        st.markdown('<div style="font-size:0.78rem;font-weight:700;color:#003DA5;'
                    'text-transform:uppercase;letter-spacing:0.06em;margin-bottom:0.3rem;">'
                    'Evaluation Ratings</div>', unsafe_allow_html=True)
        rated = [(r.get("title","")[:28], float(r["evaluation_rating"]),
                  r.get("overall_rating_label","")) for r in reports if r.get("evaluation_rating")]
        if rated:
            labels = [x[0] for x in rated]
            vals   = [x[1] for x in rated]
            clrs   = ["#22c55e" if v >= 4.5 else ("#f59e0b" if v >= 3.0 else "#ef4444") for v in vals]
            hover  = [f"<b>{x[0]}</b><br>Rating: {x[1]:.1f}/6<br>{x[2]}" for x in rated]
            fig = go.Figure(go.Bar(
                y=labels, x=vals, orientation="h",
                marker_color=clrs,
                text=[f"{v:.1f}" for v in vals],
                textposition="outside",
                hovertext=hover, hoverinfo="text",
            ))
            fig.update_layout(
                height=220, margin=dict(t=5,b=5,l=5,r=40),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(range=[0,7], showgrid=True, gridcolor="#f0f0f0",
                           title=dict(text="Rating (1–6)", font=dict(size=10))),
                yaxis=dict(tickfont=dict(size=9)),
                font=dict(size=10),
            )
            fig.add_vline(x=4.0, line_dash="dot", line_color="#9ca3af", line_width=1)
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

    with col_t:
        st.markdown('<div style="font-size:0.78rem;font-weight:700;color:#003DA5;'
                    'text-transform:uppercase;letter-spacing:0.06em;margin-bottom:0.3rem;">'
                    'Thematic Areas</div>', unsafe_allow_html=True)
        if by_thematic:
            theme_palette = ["#003DA5","#009EDB","#4CAF50","#FF9800","#9C27B0","#E91E63","#00BCD4"]
            fig = go.Figure(go.Pie(
                labels=list(by_thematic.keys()),
                values=list(by_thematic.values()),
                hole=0.45,
                marker=dict(colors=theme_palette[:len(by_thematic)],
                            line=dict(color="white", width=2)),
                textinfo="label+percent",
                textfont=dict(size=9),
                hovertemplate="<b>%{label}</b><br>%{value} report(s)<extra></extra>",
            ))
            fig.update_layout(
                height=220, margin=dict(t=5,b=5,l=5,r=5),
                paper_bgcolor="rgba(0,0,0,0)",
                showlegend=False,
            )
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

    with col_s:
        st.markdown('<div style="font-size:0.78rem;font-weight:700;color:#003DA5;'
                    'text-transform:uppercase;letter-spacing:0.06em;margin-bottom:0.3rem;">'
                    'SDG Frequency</div>', unsafe_allow_html=True)
        if by_sdg:
            top = sorted(by_sdg.items(), key=lambda x: -x[1])[:8]
            slabels = [x[0] for x in top]
            svals   = [x[1] for x in top]
            scolors = [SDG_COLORS.get(int(l.split()[1]), "#009EDB") for l in slabels]
            fig = go.Figure(go.Bar(
                y=slabels, x=svals, orientation="h",
                marker_color=scolors,
                text=svals, textposition="outside",
                hovertemplate="%{y}: %{x} report(s)<extra></extra>",
            ))
            fig.update_layout(
                height=220, margin=dict(t=5,b=5,l=5,r=30),
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                xaxis=dict(showgrid=True, gridcolor="#f0f0f0", title="Reports"),
                yaxis=dict(tickfont=dict(size=9)),
                font=dict(size=10),
            )
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    # ── Row 3: Per-report detail table ────────────────────────────────────────
    st.markdown('<div style="font-size:0.78rem;font-weight:700;color:#003DA5;'
                'text-transform:uppercase;letter-spacing:0.06em;margin-bottom:0.5rem;">'
                'Report Portfolio Overview</div>', unsafe_allow_html=True)

    def _rating_color_cell(val):
        if val is None: return "background-color:#f3f4f6;color:#9ca3af"
        v = float(val)
        if v >= 4.5: return "background-color:#dcfce7;color:#166534;font-weight:700"
        if v >= 3.0: return "background-color:#fef9c3;color:#713f12;font-weight:700"
        return "background-color:#fee2e2;color:#991b1b;font-weight:700"

    rows = []
    for r in reports:
        rating_val = r.get("evaluation_rating")
        rows.append({
            "Title": (r.get("title",""))[:55] + ("…" if len(r.get("title","")) > 55 else ""),
            "Country": r.get("country","—"),
            "Year": r.get("year","—"),
            "Thematic Area": r.get("thematic_category","—"),
            "SDGs": ", ".join(f"SDG {s}" for s in sorted(r.get("sdgs") or [])),
            "Donor": r.get("donor","—"),
            "Rating": f"{float(rating_val):.1f}/6" if rating_val else "N/A",
            "Verdict": r.get("overall_rating_label","—"),
        })

    df = pd.DataFrame(rows)

    def _style_row(row):
        rating_str = row.get("Rating","")
        try:
            v = float(rating_str.replace("/6",""))
            if v >= 4.5: bg = "#dcfce7"
            elif v >= 3.0: bg = "#fef9c3"
            else: bg = "#fee2e2"
        except Exception:
            bg = "#f9fafb"
        return [f"background-color:{bg}" if col == "Rating" else "" for col in row.index]

    styled = df.style.apply(_style_row, axis=1).set_properties(**{
        "font-size": "13px",
        "text-align": "left",
    }).set_table_styles([{
        "selector": "th",
        "props": [("background-color","#003DA5"),("color","white"),
                  ("font-size","12px"),("font-weight","600"),
                  ("text-transform","uppercase"),("letter-spacing","0.04em"),
                  ("padding","6px 10px")],
    },{
        "selector": "td",
        "props": [("padding","6px 10px"),("border-bottom","1px solid #f0f0f0")],
    }])
    st.dataframe(df, use_container_width=True, hide_index=True)

    # ── Report Infographic Generator ──────────────────────────────────────────
    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)
    st.markdown('<div style="font-size:0.78rem;font-weight:700;color:#003DA5;'
                'text-transform:uppercase;letter-spacing:0.06em;margin-bottom:0.6rem;">'
                '📊 Report Infographic</div>', unsafe_allow_html=True)

    infog_col1, infog_col2 = st.columns([3, 1])
    with infog_col1:
        report_options = {r["report_id"]: f"{r.get('year','')} — {display_title(r)}"
                         for r in reports}
        infog_rid = st.selectbox(
            "Select report",
            list(report_options.keys()),
            format_func=lambda x: report_options.get(x, x),
            key="infog_report_sel",
            label_visibility="collapsed",
        )
    with infog_col2:
        gen_btn = st.button("✨ Generate Infographic", type="primary",
                            use_container_width=True, key="infog_gen_btn")

    if gen_btn and infog_rid:
        with st.spinner("Claude is extracting insights and building your infographic…"):
            try:
                html_bytes = _build_report_infographic(infog_rid)
                st.session_state[f"infog_html_{infog_rid}"] = html_bytes
            except Exception as e:
                st.error(f"Infographic error: {e}")
                import traceback; st.code(traceback.format_exc())

    cached_html = st.session_state.get(f"infog_html_{infog_rid}")
    if cached_html:
        # Render inline preview
        st.components.v1.html(cached_html.decode("utf-8"), height=900, scrolling=True)
        dl_col, regen_col, tip_col = st.columns([2, 1, 3])
        with dl_col:
            st.download_button(
                "⬇ Download HTML",
                data=cached_html,
                file_name=f"UNIDO_{infog_rid}_Infographic.html",
                mime="text/html",
                use_container_width=True,
            )
        with regen_col:
            if st.button("↺ Regenerate", key="infog_regen", use_container_width=True):
                del st.session_state[f"infog_html_{infog_rid}"]
                st.rerun()
        with tip_col:
            st.caption("💡 Open the downloaded HTML in your browser → File → Print → Save as PDF for a print-ready copy.")


# ─────────────────────────────────────────────────────────────────────────────
# INFOGRAPHIC ENGINE
# ─────────────────────────────────────────────────────────────────────────────

def _find_pdf_path(rid: str) -> str:
    import glob, os
    base = os.path.join(os.path.dirname(__file__), "..", "data", "pdfs")
    nums = rid.split("-")[-1]
    for candidate in glob.glob(os.path.join(base, "**", f"*{nums}*.pdf"), recursive=True):
        return candidate
    return ""


def _extract_rich_infographic_data(rid: str) -> dict:
    """Extract rich infographic fields directly from the PDF. Cached in ai_extractions JSON."""
    import re as _re, json as _json, os as _os

    # Guard: pdfplumber segfaults on Streamlit Cloud — skip entirely there.
    # PILOT_METADATA already covers all 4 demo reports; corpus regex handles others.
    if _os.path.exists('/mount/src'):
        return {}

    ai_path = _os.path.join(_os.path.dirname(__file__), "..", "data",
                            "ai_extractions", f"{rid}.json")
    cached_ai = {}
    if _os.path.exists(ai_path):
        try:
            with open(ai_path, "r", encoding="utf-8") as f:
                cached_ai = _json.load(f)
            if "infographic_data" in cached_ai:
                return cached_ai["infographic_data"]
        except Exception:
            pass

    pdf_path = _find_pdf_path(rid)
    if not pdf_path:
        return {}

    try:
        import pdfplumber as _plumber
        with _plumber.open(pdf_path) as pdf:
            full_text = "\n".join((p.extract_text() or "") for p in pdf.pages)
    except Exception:
        return {}

    result: dict = {}

    # GEF ID
    for pat in [
        r'GEF[-\s]?ID[-:\s]+(\d{4,6})',
        r'GEF%20ID-(\d{4,6})',
        r'GEF\s+project\s+(?:id|no\.?)\s*[:\-]?\s*(\d{4,6})',
    ]:
        m = _re.search(pat, full_text, _re.IGNORECASE)
        if m:
            result["gef_id"] = m.group(1)
            break

    # Planned duration in months
    for pat in [
        r'(?:approved|designed?|planned).*?(?:for a period of|over|duration of)\s+(\d+)\s*months?',
        r'(?:period of|over)\s+(\d+)\s*months?\s*(?:of\s+)?(?:implementation|project)',
        r'(\d+)[\s-]month\s+(?:implementation\s+)?(?:project|period)',
    ]:
        m = _re.search(pat, full_text, _re.IGNORECASE)
        if m:
            v = int(m.group(1))
            if 6 <= v <= 120:
                result["planned_months"] = v
                break

    # Start date
    MONTHS = {'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
              'july':7,'august':8,'september':9,'october':10,'november':11,'december':12}
    for pat in [
        r'Implementation\s+Start\s+Date\s+(\d{1,2})\s+(\w+)\s+(\d{4})',
        r'(?:start of|officially marked.*?start).{0,80}?(\d{1,2})\s+(\w+)\s+(\d{4})',
    ]:
        m = _re.search(pat, full_text, _re.IGNORECASE)
        if m:
            mo_name = m.group(2).lower()
            yr = int(m.group(3))
            mo = MONTHS.get(mo_name, 0)
            if 2000 <= yr <= 2030 and mo:
                result["start_year"]  = yr
                result["start_month"] = mo
                break
    if "start_year" not in result:
        m = _re.search(r'(?:started?|commenced?|initiated?|launched?)\s+in\s+(20\d\d)',
                       full_text, _re.IGNORECASE)
        if m:
            result["start_year"] = int(m.group(1))

    # Planned end date
    for pat in [
        r'original\s+(?:closure|completion|end)\s+date.*?(?:in|of|:)\s+(\w+\s+\d{4}|\d{4})',
        r'(?:planned|original)\s+(?:end|closure|completion).*?(\w+\s+\d{4}|\d{4})',
    ]:
        m = _re.search(pat, full_text, _re.IGNORECASE)
        if m:
            raw = m.group(1).strip()
            yrm = _re.search(r'(\d{4})', raw)
            if yrm:
                result["planned_end_year"] = int(yrm.group(1))
            break
    if "planned_end_year" not in result and "start_year" in result and "planned_months" in result:
        result["planned_end_year"] = result["start_year"] + round(result["planned_months"] / 12)

    # Actual end date
    for pat in [
        r'extended.*?until.*?(?:end\s+of\s+)?(\w+\s+\d{4}|\d{4})',
        r'(?:project\s+)?(?:closed?|completed?|concluded?)\s+in\s+(\w+\s+\d{4}|\d{4})',
    ]:
        m = _re.search(pat, full_text, _re.IGNORECASE)
        if m:
            raw = m.group(1).strip()
            yrm = _re.search(r'(\d{4})', raw)
            if yrm:
                result["actual_end_year"] = int(yrm.group(1))
            break

    # Compute actual_months & delay
    if "start_year" in result and "actual_end_year" in result:
        sm = result.get("start_month", 6)
        actual_months = (result["actual_end_year"] - result["start_year"]) * 12 + (12 - sm)
        result["actual_months"] = round(actual_months)
    if "planned_months" in result and "actual_months" in result:
        dm = result["actual_months"] - result["planned_months"]
        if dm > 0:
            result["has_delay"]    = True
            result["delay_months"] = dm
            result["overrun_pct"]  = round(dm / result["planned_months"] * 100)
        else:
            result["has_delay"] = False
    if "start_year" in result and "actual_end_year" in result:
        result["duration_str"] = f"{result['start_year']} – {result['actual_end_year']}"

    # Year-by-year disbursement table
    yh = _re.search(r'(?:Component|Activity|Year)\s+((?:20\d\d\s+){2,})', full_text, _re.IGNORECASE)
    tot = _re.search(r'^Total\s+([\d,.\s]+)$', full_text, _re.IGNORECASE | _re.MULTILINE)
    if yh and tot:
        raw_years  = _re.findall(r'20\d\d', yh.group(1))
        raw_amts   = [x for x in _re.split(r'\s+', tot.group(1).strip()) if x]
        amounts = []
        for a in raw_amts:
            try:
                amounts.append(float(a.replace(',', '')))
            except ValueError:
                pass
        if len(amounts) == len(raw_years) + 1:
            amounts = amounts[:len(raw_years)]
        if raw_years and len(raw_years) == len(amounts) and len(raw_years) >= 2:
            result["disbursement_by_year"] = {
                yr: round(am) for yr, am in zip(raw_years, amounts) if am > 0
            }

    # ── Stakeholder gender table ─────────────────────────────────────────────
    # Looks for a table like: Name | Total employees | Female | % women
    # Appears in UNIDO TEs under "Gender analysis" or "Table X: Gender"
    stakeholders_rich = []
    _ROLE_HINTS = [
        (r'national\s+(?:implement|coordinat|partner|focal)',   "National implementing partner"),
        (r'implement.*?partner',                                "Implementing partner"),
        (r'refriger|RAC|air.?condit',                          "Enterprise: refrigeration"),
        (r'foam|PU|polyurethane',                              "Enterprise: foam sector"),
        (r'chemical|methyl',                                    "Enterprise: chemicals"),
        (r'training\s+cent',                                    "Training centre"),
        (r'government|minist|ministry',                         "Government counterpart"),
        (r'manufactur|convers',                                 "Manufacturing beneficiary"),
        (r'cleaner\s+product|RECP|NCPC',                       "Cleaner production centre"),
        (r'enterprise|company|compan',                          "Project enterprise"),
    ]

    # Strategy 1: Find "% of women employed" table header then parse rows
    gender_block = None
    for pat in [
        r'% of women employed(.*?)(?:\n{3,}|\Z)',
        r'(?:Gender analysis|Table \d+[.:][^\n]*gender)(.*?)(?:\n{3,}|\Z)',
        r'(?:gender.{0,60}?mainstreaming|gender.{0,60}?analysis)(.*?)(?:\n{3,}|\Z)',
    ]:
        m = _re.search(pat, full_text, _re.IGNORECASE | _re.DOTALL)
        if m:
            gender_block = m.group(1)
            break

    if gender_block:
        lines = [l.strip() for l in gender_block.split('\n') if l.strip()]
        i = 0
        while i < len(lines):
            name = lines[i]
            # Skip header-like lines and pure numbers
            if (len(name) < 3 or name.isdigit() or
                    _re.match(r'^(name|stakeholder|organisation|total|female|women|%|no\.)', name, _re.I)):
                i += 1
                continue
            # Look ahead for: total_employees female_employees pct
            # Pattern A: next 3 tokens are numbers (total, female, pct)
            nums_found = []
            for j in range(i+1, min(i+6, len(lines))):
                if _re.match(r'^\d+[\d,.\s\(\)]*$', lines[j]):
                    nums_found.append((j, float(_re.sub(r'[,\s\(\)]','',lines[j]) or '0')))
                elif nums_found:
                    break
            if len(nums_found) >= 2:
                # Last number is pct if ≤100, else look for explicit pct
                vals = [v for _,v in nums_found]
                pct = None
                if vals[-1] <= 100:
                    pct = round(vals[-1])
                elif len(vals) >= 2 and vals[0] > 0:
                    pct = round(vals[1] / vals[0] * 100)
                if pct is not None and 0 <= pct <= 100:
                    # Guess role from context around name in full text
                    role = "Project stakeholder"
                    ctx_snip = full_text[max(0, full_text.find(name)-200):full_text.find(name)+200]
                    for rpat, rlabel in _ROLE_HINTS:
                        if _re.search(rpat, ctx_snip, _re.I):
                            role = rlabel
                            break
                    stakeholders_rich.append({"name": name, "role": role, "pct_women": pct})
                    i = nums_found[-1][0] + 1
                    continue
            i += 1

    # Strategy 2: inline "X% women" pattern near company names
    if not stakeholders_rich:
        for m in _re.finditer(
            r'([A-Z][A-Za-z0-9\s\./&\-]{2,40})\s*[:\-–]?\s*.*?(\d{1,3})\s*%\s*(?:women|female)',
            full_text, _re.IGNORECASE
        ):
            name = m.group(1).strip()
            pct  = int(m.group(2))
            if 0 <= pct <= 100 and len(name) > 3:
                role = "Project stakeholder"
                ctx_snip = full_text[max(0,m.start()-150):m.end()+150]
                for rpat, rlabel in _ROLE_HINTS:
                    if _re.search(rpat, ctx_snip, _re.I):
                        role = rlabel
                        break
                if not any(s["name"] == name for s in stakeholders_rich):
                    stakeholders_rich.append({"name": name, "role": role, "pct_women": pct})
        if len(stakeholders_rich) > 8:
            stakeholders_rich = stakeholders_rich[:8]

    if stakeholders_rich:
        result["stakeholders_rich"] = stakeholders_rich

    # ── Delay causes ─────────────────────────────────────────────────────────
    delay_sents = []
    for sent in _re.split(r'(?<=[.!?])\s+', full_text):
        s = sent.strip()
        if (_re.search(r'\b(?:delay|extend|postpone|behind\s+schedule|covid|pandemic|overrun|late)\b',
                       s, _re.I) and 40 < len(s) < 300 and s not in delay_sents):
            delay_sents.append(s)
        if len(delay_sents) >= 3:
            break
    if delay_sents:
        result["delay_causes"] = delay_sents

    # Cache
    if result and cached_ai:
        try:
            cached_ai["infographic_data"] = result
            with open(ai_path, "w", encoding="utf-8") as f:
                _json.dump(cached_ai, f, ensure_ascii=False, indent=2)
        except Exception:
            pass

    return result


def _parse_report_data(rid: str) -> dict:
    """No-LLM parser. Extracts infographic data from ai_extractions + extracted_sections
    + PILOT_METADATA + PDF (for non-pilot reports). Results cached."""
    import re

    ai    = _load_ai_extraction(rid)
    sec   = _load_sections_local(rid)
    pilot = PILOT_METADATA.get(rid, {})
    ctx   = ai.get("context", {})

    title      = ctx.get("title")       or pilot.get("title", rid)
    year       = ctx.get("year")        or pilot.get("year")
    country    = ctx.get("country")     or pilot.get("country", "")
    region     = ctx.get("region")      or pilot.get("region", "")
    rtype      = ctx.get("report_type") or pilot.get("report_type", "Terminal Evaluation")
    donor      = ctx.get("donor")       or pilot.get("donor", "")
    project_id = ctx.get("project_id")  or pilot.get("project_id", "")
    thematic   = ai.get("primary_thematic_area") or pilot.get("thematic_category", "")
    budget_usd = ctx.get("budget_usd")  or pilot.get("budget_usd")
    try:
        year = int(year) if year else None
    except (ValueError, TypeError):
        year = None
    try:
        budget_usd = float(budget_usd) if budget_usd else None
    except (ValueError, TypeError):
        budget_usd = None

    secs_data = sec.get("sections", {})
    corpus = "\n\n".join(p for p in [
        ai.get("executive_summary", ""),
        secs_data.get("conclusions", "") or "",
        secs_data.get("findings", "") or secs_data.get("results", "") or "",
    ] if p and len(p) > 50)

    # Duration via regex on corpus
    month_re = re.compile(r'(\d+)\s*[-–]?\s*month(?:s)?', re.IGNORECASE)
    month_vals = [int(m.group(1)) for m in month_re.finditer(corpus) if 6 <= int(m.group(1)) <= 240]
    year_dur_re = re.compile(
        r'(\d+)\s+year[s]?\s+(?:implementation\s+)?(?:period|duration|project)', re.IGNORECASE)
    year_dur_vals = [int(m.group(1)) * 12 for m in year_dur_re.finditer(corpus)
                     if 1 <= int(m.group(1)) <= 20]
    all_durations = sorted(set(month_vals + year_dur_vals))

    ext_re = re.compile(r'(?:extended?|extension)\s+(?:by|of)\s+(\d+)\s*month', re.IGNORECASE)
    ext_match = ext_re.search(corpus)
    ext_months = int(ext_match.group(1)) if ext_match else 0

    planned_months = actual_months = None
    if len(all_durations) >= 2:
        planned_months = all_durations[0]
        actual_months  = all_durations[-1]
    elif len(all_durations) == 1:
        planned_months = all_durations[0]
        actual_months  = planned_months + ext_months if ext_months else planned_months

    # Start/end year via regex on corpus
    start_year = end_year = None
    range_re = re.compile(r'\b(20[01]\d)\s*(?:[-–]|to|through)\s*(20[012]\d)\b')
    for m in range_re.finditer(corpus):
        sy, ey = int(m.group(1)), int(m.group(2))
        if ey > sy and (ey - sy) <= 20:
            start_year, end_year = sy, ey
            break
    if not start_year:
        sp = re.search(r'(?:commenced?|started?|initiated?|launched?|began)\s+in\s+(20[01]\d)',
                       corpus, re.IGNORECASE)
        if sp: start_year = int(sp.group(1))
    if not end_year:
        ep = re.search(r'(?:ended?|completed?|closed?|concluded?)\s+in\s+(20[012]\d)',
                       corpus, re.IGNORECASE)
        if ep: end_year = int(ep.group(1))
    if not start_year and year and planned_months:
        start_year = year - round(planned_months / 12)
    if not end_year and year:
        end_year = year

    planned_end_year = actual_end_year = None
    if start_year and planned_months:
        planned_end_year = start_year + round(planned_months / 12)
    elif end_year:
        planned_end_year = end_year
    if start_year and actual_months:
        actual_end_year = start_year + round(actual_months / 12)
    elif end_year:
        actual_end_year = end_year

    has_delay    = bool(planned_months and actual_months and actual_months > planned_months)
    delay_months = (actual_months - planned_months) if has_delay else 0

    # Financial via regex
    def _extract_amounts(text):
        amounts = []
        for pat, mult in [
            (r'USD\s+([\d,]+(?:\.\d+)?)\s*million', 1e6),
            (r'US\$\s*([\d,]+(?:\.\d+)?)\s*million', 1e6),
            (r'\$([\d,]+(?:\.\d+)?)\s*million', 1e6),
            (r'USD\s*([\d,]+)', 1.0),
            (r'US\$\s*([\d,]+)', 1.0),
            (r'\$\s*([\d,]+)', 1.0),
        ]:
            for m in re.finditer(pat, text, re.IGNORECASE):
                try:
                    v = float(m.group(1).replace(',', '')) * mult
                    if v >= 10_000:
                        amounts.append(round(v))
                except ValueError:
                    pass
        return amounts

    raw_amounts = sorted(set(_extract_amounts(corpus)), reverse=True)
    deduped = []
    for a in raw_amounts:
        if not deduped or abs(a - deduped[-1]) / max(deduped[-1], 1) > 0.05:
            deduped.append(a)

    expenditure_usd = None
    if budget_usd:
        exp_candidates = [a for a in deduped if a < budget_usd * 1.05 and abs(a - budget_usd) / budget_usd > 0.03]
        if exp_candidates:
            expenditure_usd = min(exp_candidates, key=lambda x: abs(x - budget_usd * 0.9))
    else:
        if deduped:
            budget_usd = deduped[0]
        if len(deduped) >= 2:
            expenditure_usd = deduped[1]

    util_rate = None
    um = re.search(r'(\d+(?:\.\d+)?)\s*%\s+(?:of\s+)?(?:budget|funds?|allocation|disburs|utiliz)',
                   corpus, re.IGNORECASE)
    if um:
        util_rate = float(um.group(1))
    elif budget_usd and expenditure_usd:
        util_rate = round(expenditure_usd / budget_usd * 100, 1)

    # Stakeholders via keyword scan
    STAKEHOLDER_TERMS = [
        ("National Government",  r'\b(?:government|ministry|ministries|national\s+authority)\b'),
        ("Private Sector",       r'\bprivate\s+sector\b'),
        ("Enterprises / SMEs",   r'\b(?:enterprise|sme|company|compan|firm|industr)\b'),
        ("Civil Society / NGOs", r'\b(?:ngo|civil\s+society|association|federation)\b'),
        ("Local Communities",    r'\b(?:communit|local|rural|village|household)\b'),
        ("Academic / Research",  r'\b(?:universit|research|academic|institut)\b'),
        ("International Orgs",   r'\b(?:unep|undp|worldbank|world\s+bank|bilateral)\b'),
    ]
    stakeholders = [name for name, pat in STAKEHOLDER_TERMS
                    if re.search(pat, corpus, re.IGNORECASE)]

    # Gender
    gender_count = len(re.findall(r'\b(?:gender|women|female|girl)\b', corpus, re.IGNORECASE))
    gender_pct = None
    gm = re.search(r'(\d+(?:\.\d+)?)\s*%\s+(?:women|female|girl)', corpus, re.IGNORECASE)
    if gm:
        gender_pct = float(gm.group(1))
    if gender_count >= 10:
        gender_rating, gender_color = "Mainstreamed", "#2f7a6f"
    elif gender_count >= 3:
        gender_rating, gender_color = "Partially Mainstreamed", "#c08a34"
    else:
        gender_rating, gender_color = "Not Mainstreamed", "#a3492f"
    gender_note = None
    for sent in re.split(r'(?<=[.!?])\s+', corpus):
        if re.search(r'\b(?:gender|women|female)\b', sent, re.IGNORECASE) and 30 < len(sent) < 300:
            gender_note = sent.strip()
            break

    # Delay causes
    delay_causes = []
    for sent in re.split(r'(?<=[.!?])\s+', corpus):
        if re.search(r'\b(?:delay|extend|postpone|behind\s+schedule|covid|pandemic)\b',
                     sent, re.IGNORECASE):
            s = sent.strip()
            if 20 < len(s) < 200 and s not in delay_causes:
                delay_causes.append(s)
    delay_causes = delay_causes[:3]

    # GEF ID
    gef_id = None
    for gpat in [r'GEF\s+(?:Project\s+)?(?:ID|No\.?)\s*:?\s*([\d]{3,6})',
                 r'\bGEF[/-]([\d]{3,6})\b']:
        gm2 = re.search(gpat, corpus, re.IGNORECASE)
        if gm2:
            gef_id = gm2.group(1)
            break

    # Overall rating
    overall_rating_label = pilot.get("overall_rating_label", "")
    if not overall_rating_label:
        rating_val = ctx.get("evaluation_rating") or pilot.get("evaluation_rating")
        if rating_val:
            try:
                rv = float(rating_val)
                if rv >= 5.5:   overall_rating_label = "Highly Satisfactory"
                elif rv >= 4.5: overall_rating_label = "Satisfactory"
                elif rv >= 3.5: overall_rating_label = "Moderately Satisfactory"
                elif rv >= 2.5: overall_rating_label = "Moderately Unsatisfactory"
                elif rv >= 1.5: overall_rating_label = "Unsatisfactory"
                else:           overall_rating_label = "Highly Unsatisfactory"
            except (ValueError, TypeError):
                pass

    # PDF-based rich extraction (for non-pilot reports, or any report missing rich fields)
    if not pilot.get("disbursement_by_year") and not pilot.get("stakeholders_rich"):
        try:
            pdf_rich = _extract_rich_infographic_data(rid)
        except Exception:
            pdf_rich = {}
    else:
        pdf_rich = {}

    def _pf(field, current):
        return pdf_rich.get(field) if current is None and field in pdf_rich else current

    planned_months   = _pf("planned_months",   planned_months)
    actual_months    = _pf("actual_months",     actual_months)
    start_year       = _pf("start_year",        start_year)
    planned_end_year = _pf("planned_end_year",  planned_end_year)
    actual_end_year  = _pf("actual_end_year",   actual_end_year)
    gef_id           = _pf("gef_id",            gef_id)
    if not delay_causes and pdf_rich.get("delay_causes"):
        delay_causes = pdf_rich["delay_causes"]

    disbursement_by_year = pdf_rich.get("disbursement_by_year")
    stakeholders_rich    = pdf_rich.get("stakeholders_rich")
    cofinancing_labels = cofinancing_planned = cofinancing_actual = None

    if pdf_rich.get("has_delay") is not None and not has_delay:
        has_delay    = pdf_rich["has_delay"]
        delay_months = pdf_rich.get("delay_months", 0)
    overrun_pct_pdf = pdf_rich.get("overrun_pct")
    if pdf_rich.get("duration_str") and not (start_year and actual_end_year):
        pass  # will be set below

    # Recompute delay after PDF merge
    if planned_months and actual_months and actual_months > planned_months:
        has_delay    = True
        delay_months = actual_months - planned_months

    # PILOT_METADATA rich overrides (highest priority — pre-computed correct values)
    if pilot.get("planned_months"):
        planned_months = pilot["planned_months"]
    if pilot.get("actual_months"):
        actual_months = pilot["actual_months"]
    if "delay_months" in pilot:
        delay_months = pilot["delay_months"]
    if "has_delay" in pilot:
        has_delay = pilot["has_delay"]
    if pilot.get("start_year"):
        start_year = pilot["start_year"]
    if pilot.get("planned_end_year"):
        planned_end_year = pilot["planned_end_year"]
    if pilot.get("actual_end_year"):
        actual_end_year = pilot["actual_end_year"]
    if pilot.get("gef_id") and not gef_id:
        gef_id = pilot["gef_id"]
    if pilot.get("util_rate") is not None:
        util_rate = pilot["util_rate"]
    if pilot.get("disbursement_by_year"):
        disbursement_by_year = pilot["disbursement_by_year"]
    if pilot.get("cofinancing_labels"):
        cofinancing_labels  = pilot["cofinancing_labels"]
        cofinancing_planned = pilot.get("cofinancing_planned", [])
        cofinancing_actual  = pilot.get("cofinancing_actual",  [])
    if pilot.get("stakeholders_rich"):
        stakeholders_rich = pilot["stakeholders_rich"]
    if pilot.get("gender_rating"):
        gender_rating = pilot["gender_rating"]
        gender_color  = {"Mainstreamed":"#2f7a6f","Partially Mainstreamed":"#c08a34"}.get(gender_rating,"#a3492f")
    if pilot.get("gender_note"):
        gender_note = pilot["gender_note"]
    if pilot.get("delay_causes"):
        delay_causes = pilot["delay_causes"]
    if pilot.get("timeline_events"):
        timeline_events = pilot["timeline_events"]
    else:
        timeline_events = None
    if pilot.get("budget_usd") and not budget_usd:
        budget_usd = float(pilot["budget_usd"])

    # Overrun pct
    overrun_pct = pilot.get("overrun_pct") or overrun_pct_pdf
    if overrun_pct is None and has_delay and planned_months:
        overrun_pct = round(delay_months / planned_months * 100)

    # Duration string
    duration_str = pilot.get("duration_str", "")
    if not duration_str:
        if pdf_rich.get("duration_str"):
            duration_str = pdf_rich["duration_str"]
        elif start_year and actual_end_year:
            duration_str = f"{start_year} – {actual_end_year}"
        elif start_year and planned_end_year:
            duration_str = f"{start_year} – {planned_end_year}"

    return {
        "rid": rid, "title": title, "year": year,
        "country": country, "region": region,
        "report_type": rtype, "thematic": thematic,
        "donor": donor, "project_id": project_id,
        "gef_id": gef_id,
        "budget_usd": budget_usd, "expenditure_usd": expenditure_usd,
        "util_rate": util_rate,
        "planned_months": planned_months, "actual_months": actual_months,
        "start_year": start_year, "end_year": end_year,
        "planned_end_year": planned_end_year, "actual_end_year": actual_end_year,
        "has_delay": has_delay, "delay_months": delay_months,
        "delay_causes": delay_causes,
        "stakeholders": stakeholders, "stakeholders_rich": stakeholders_rich,
        "gender_rating": gender_rating, "gender_color": gender_color,
        "gender_count": gender_count, "gender_pct": gender_pct, "gender_note": gender_note,
        "overall_rating_label": overall_rating_label,
        "overrun_pct": overrun_pct, "duration_str": duration_str,
        "timeline_events": timeline_events,
        "disbursement_by_year": disbursement_by_year,
        "cofinancing_labels": cofinancing_labels,
        "cofinancing_planned": cofinancing_planned,
        "cofinancing_actual":  cofinancing_actual,
    }


def _render_infographic_html(d: dict) -> str:
    """Render the 5-section A4 infographic matching the UNIDO IEU standardized template."""
    import json as _json

    def _esc(s):
        return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
    def _fmt_m(m):
        if m is None: return "N/A"
        yrs, mos = divmod(int(m), 12)
        if mos == 0: return f"{yrs} yr{'s' if yrs != 1 else ''}"
        if yrs == 0: return f"{mos} mo"
        return f"{yrs} yr{'s' if yrs != 1 else ''} {mos} mo"
    def _fmt_usd(v):
        if v is None: return "N/A"
        if v >= 1_000_000: return f"${v/1_000_000:.2f}M"
        if v >= 1_000: return f"${v/1_000:.0f}K"
        return f"${v:.0f}"

    # Timeline
    sy=d.get("start_year"); pey=d.get("planned_end_year")
    aey=d.get("actual_end_year"); y=d.get("year")
    if d.get("timeline_events"):
        raw_events = d["timeline_events"]
        # PILOT_METADATA events have yr/lab/type but no pos — compute pos now
        if raw_events and "pos" not in raw_events[0]:
            yrs = [e["yr"] for e in raw_events]
            tl_min=min(yrs); tl_max=max(yrs); tl_span=max(tl_max-tl_min,1)
            raw_events = [dict(e, pos=round((e["yr"]-tl_min)/tl_span*92+4,1),
                               yr=str(e["yr"])) for e in raw_events]
        tl_events_json = _json.dumps(raw_events)
        has_timeline   = True
    else:
        tl_events_json = "[]"; has_timeline = bool(sy and (pey or aey))
        if has_timeline:
            all_yrs = [x for x in [sy,pey,aey,y] if x]
            tl_min=min(all_yrs); tl_max=max(all_yrs); tl_span=max(tl_max-tl_min,1)
            def _pos(yr): return round((yr-tl_min)/tl_span*92+4,1)
            events=[]
            if sy: events.append({"yr":str(sy),"pos":_pos(sy),"lab":"Project Start","type":"plan"})
            if sy and pey and (pey-sy)>1:
                mid=round(sy+(pey-sy)*0.45)
                events.append({"yr":str(mid),"pos":_pos(mid),"lab":"Mid-Term Review","type":"actual"})
            if pey and d.get("has_delay"):
                events.append({"yr":f"{pey}*","pos":_pos(pey),"lab":"Planned Closure","type":"plan"})
            if aey:
                lbl=f"Actual Closure (+{d['delay_months']}mo)" if d.get("has_delay") else "Project Closure"
                events.append({"yr":str(aey),"pos":_pos(aey),"lab":lbl,"type":"actual"})
            if y and y!=aey:
                events.append({"yr":str(y),"pos":_pos(y),"lab":"Terminal Evaluation","type":"actual"})
            tl_events_json=_json.dumps(events)

    # Gantt
    pm=d.get("planned_months"); am=d.get("actual_months")
    has_gantt=bool(d.get("has_delay") and pm and am)
    p_pct=ov_pct=0
    if has_gantt:
        p_pct=round(pm/am*100,1); ov_pct=round(100-p_pct,1)

    # Financial
    bv=d.get("budget_usd"); ev=d.get("expenditure_usd")
    ur=d.get("util_rate"); donor=d.get("donor","")
    has_financial=bool(bv)
    dby=d.get("disbursement_by_year")
    cofin_labels=d.get("cofinancing_labels")
    cofin_planned=d.get("cofinancing_planned")
    cofin_actual=d.get("cofinancing_actual")

    if dby:
        spend_chart_title=f"{_esc(donor)} Disbursement by Year (US$)" if donor else "Disbursement by Year (US$)"
        spend_chart_js=f"""new Chart(document.getElementById('spendChart'),{{
    type:'bar',data:{{labels:{_json.dumps(list(dby.keys()))},
    datasets:[{{data:{_json.dumps(list(dby.values()))},backgroundColor:'#173a58',borderRadius:2}}]}},
    options:{{plugins:{{legend:{{display:false}}}},scales:{{
      y:{{beginAtZero:true,ticks:{{callback:function(v){{return v>=1e6?'$'+(v/1e6).toFixed(1)+'M':'$'+(v/1000).toFixed(0)+'K';}}}}}},
      x:{{grid:{{display:false}}}}}}
    }}}});"""
    elif bv:
        spend_chart_title=f"{_esc(donor)} Grant" if donor else "Budget"
        fl=_json.dumps([f"{donor} Grant" if donor else "Budget"]+([" Expenditure"] if ev else []))
        fv=_json.dumps([round(bv)]+([round(ev)] if ev else []))
        fc=_json.dumps(["#173a58"]+([" #2f7a6f"] if ev else []))
        spend_chart_js=f"""new Chart(document.getElementById('spendChart'),{{
    type:'bar',data:{{labels:{fl},datasets:[{{data:{fv},backgroundColor:{fc},borderRadius:3}}]}},
    options:{{plugins:{{legend:{{display:false}}}},scales:{{
      y:{{beginAtZero:true,ticks:{{callback:function(v){{return v>=1e6?'$'+(v/1e6).toFixed(1)+'M':'$'+(v/1000).toFixed(0)+'K';}}}}}},
      x:{{grid:{{display:false}}}}}}}}
    }}}});"""
    else:
        spend_chart_title="Budget Allocation"; spend_chart_js=""

    if cofin_labels and cofin_planned is not None and cofin_actual is not None:
        cofin_chart_title="Co-financing by Source: Planned vs. Actual"
        cofin_chart_js=f"""new Chart(document.getElementById('cofinChart'),{{
    type:'bar',data:{{labels:{_json.dumps(cofin_labels)},
    datasets:[{{label:'Planned',data:{_json.dumps(cofin_planned)},backgroundColor:'#c08a34'}},
              {{label:'Actual', data:{_json.dumps(cofin_actual)}, backgroundColor:'#2f7a6f'}}]}},
    options:{{plugins:{{legend:{{position:'bottom',labels:{{boxWidth:10,font:{{size:9}}}}}}}},
      scales:{{y:{{beginAtZero:true,ticks:{{callback:function(v){{return v>=1e6?'$'+(v/1e6).toFixed(1)+'M':'$'+(v/1000).toFixed(0)+'K';}}}}}},x:{{grid:{{display:false}}}}}}
    }}}});"""
    elif ur is not None:
        cofin_chart_title="Budget Utilisation"
        uc2="#2f7a6f" if ur>=85 else "#c08a34" if ur>=60 else "#a3492f"
        cofin_chart_js=f"""new Chart(document.getElementById('cofinChart'),{{
    type:'bar',data:{{labels:['Budget Utilisation'],datasets:[{{data:[{round(ur,1)}],backgroundColor:['{uc2}'],label:'%'}}]}},
    options:{{indexAxis:'y',plugins:{{legend:{{display:false}}}},scales:{{x:{{beginAtZero:true,max:100}},y:{{grid:{{display:false}}}}}}
    }}}});"""
    else:
        cofin_chart_title="Co-financing"; cofin_chart_js=""

    finbar_html=""
    if bv:
        lab1=f"{_esc(donor)} Grant" if donor else "Budget"
        finbar_html+=(f'<div class="finbar"><div class="finlabel">{lab1}</div>'
                     f'<div class="finbg"><div class="finfill" style="width:100%;"></div></div>'
                     f'<div class="finval">{_esc(_fmt_usd(bv))}</div></div>')
    if ev and bv:
        ep=min(round(ev/bv*100),100)
        finbar_html+=(f'<div class="finbar"><div class="finlabel">Expenditure</div>'
                     f'<div class="finbg"><div class="finfill" style="width:{ep}%;background:linear-gradient(90deg,#c08a34,#dba75a);"></div></div>'
                     f'<div class="finval">{_esc(_fmt_usd(ev))}</div></div>')
    if ur is not None:
        uc="var(--ozone)" if ur>=85 else "var(--amber)" if ur>=60 else "var(--brick)"
        finbar_html+=(f'<div class="finbar"><div class="finlabel">Utilisation</div>'
                     f'<div class="finbg"><div class="finfill" style="width:{min(ur,100):.0f}%;background:{uc};"></div></div>'
                     f'<div class="finval">{ur:.1f}%</div></div>')

    # Stakeholder table
    stakeholders_rich=d.get("stakeholders_rich")
    stake_rows=""
    if stakeholders_rich:
        for s in stakeholders_rich:
            pct=s.get("pct_women",0) or 0
            stake_rows+=(f'<tr><td>{_esc(s.get("name",""))}</td><td>{_esc(s.get("role",""))}</td>'
                        f'<td><div class="gbar"><div class="gbar-bg"><div class="gbar-fill" style="width:{pct}%;"></div></div>'
                        f'<div class="gbar-val">{pct}%</div></div></td></tr>')
    else:
        for s in (d.get("stakeholders") or []):
            stake_rows+=(f'<tr><td>{_esc(s)}</td><td>Beneficiary / Partner</td>'
                        f'<td><div class="gbar"><div class="gbar-bg"><div class="gbar-fill" style="width:0%;"></div></div>'
                        f'<div class="gbar-val">—</div></div></td></tr>')
    if not stake_rows:
        stake_rows='<tr><td colspan="3" style="color:var(--ink-soft);font-style:italic;padding:8px;">Stakeholder data not available.</td></tr>'
    stk_tag="% women employed / in executive roles, by stakeholder" if stakeholders_rich else "involvement by stakeholder group"

    g_note=d.get("gender_note") or "No formal gender strategy identified in this evaluation report."
    g_rating=d.get("gender_rating","Not Assessed")

    # ID bar
    idbar_parts=[f'<div><span>UNIDO ID</span><b>{_esc(d.get("project_id") or "—")}</b></div>']
    if d.get("gef_id"):
        idbar_parts.append(f'<div><span>GEF ID</span><b>{_esc(d["gef_id"])}</b></div>')
    idbar_parts.append(f'<div><span>Country</span><b>{_esc(d.get("country") or "—")}</b></div>')
    if d.get("duration_str"):
        idbar_parts.append(f'<div><span>Duration</span><b>{_esc(d["duration_str"])}</b></div>')
    if d.get("year"):
        idbar_parts.append(f'<div><span>TE Date</span><b>{_esc(str(d["year"]))}</b></div>')
    idbar_html="".join(idbar_parts)

    # Verdict strip
    overall=d.get("overall_rating_label","")
    verdict_extra=""
    if d.get("has_delay") and d.get("delay_months"):
        pct_s=f" (+{d['overrun_pct']}%)" if d.get("overrun_pct") else ""
        verdict_extra=f"&nbsp;·&nbsp; {d['delay_months']}-month schedule overrun{pct_s} against the original plan"
    verdict_html=(f'<div class="verdict">Overall objective achievement: '
                 f'<span class="big">{_esc(overall)}</span>{verdict_extra}</div>') if overall else ""

    n4=4 if has_gantt else 3; n5=5 if has_gantt else 4
    overrun_val=(f"+{d['overrun_pct']}%" if d.get("overrun_pct")
                 else f"+{d['delay_months']}mo" if d.get("delay_months") else "On Schedule")
    overrun_cls="meta warn" if d.get("has_delay") else "meta"
    budget_disp=f"{_fmt_usd(bv)} ({_esc(donor)})" if donor and bv else _fmt_usd(bv)
    causes_html=""
    if has_gantt and d.get("delay_causes"):
        dc = d["delay_causes"]
        # delay_causes may be a plain string or a list of strings
        if isinstance(dc, str):
            dc_items = [s.strip() for s in dc.split(".") if len(s.strip()) > 20]
            if not dc_items:
                dc_items = [dc]
        else:
            dc_items = list(dc)
        items="".join(f"<li>{_esc(c)}</li>" for c in dc_items)
        causes_html=(f'<div class="delay-causes"><b style="font-size:10.5px;color:var(--unblue-deep);">'
                    f'Primary delay driver</b><ul>{items}</ul></div>')
    t_esc=_esc(d.get("title","")); rid_esc=_esc(d.get("rid",""))
    donor_line=("Funded by <b>"+_esc(donor)+"</b>. " if donor else "")+\
               ("Implemented in <b>"+_esc(d.get("country",""))+"</b>." if d.get("country") else "")

    # Pre-compute financial JS block to avoid nested f-string (Python < 3.12 limitation)
    if has_financial:
        financial_js_block = "(function(){\n  if(typeof Chart==='undefined')return;\n  " + spend_chart_js + "\n  " + cofin_chart_js + "\n})();"
    else:
        financial_js_block = ""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Evaluation Infographic — {rid_esc}</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.0/chart.umd.min.js"></script>
<style>
  :root{{--ink:#132436;--ink-soft:#5a6b7a;--paper:#faf7f0;--panel:#ffffff;
    --unblue:#173a58;--unblue-deep:#0d2436;--ozone:#2f7a6f;--amber:#c08a34;--brick:#a3492f;
    --line:#e2dcc9;--serif:"Iowan Old Style","Palatino Linotype",Georgia,serif;
    --sans:"Segoe UI",Arial,sans-serif;}}
  *{{box-sizing:border-box;}}
  body{{margin:0;background:#ded7c4;font-family:var(--sans);color:var(--ink);}}
  .page{{width:794px;margin:24px auto;background:var(--paper);box-shadow:0 10px 40px rgba(0,0,0,.25);overflow:hidden;padding:0 0 22px;}}
  .body{{padding:20px 34px 0;}}
  .section-label{{font-family:var(--serif);font-size:13px;color:var(--unblue-deep);text-transform:uppercase;letter-spacing:.08em;border-bottom:1.5px solid var(--unblue);padding-bottom:5px;margin:28px 0 12px;display:flex;justify-content:space-between;align-items:baseline;}}
  .section-label .tag{{font-size:9.5px;color:var(--ink-soft);text-transform:none;letter-spacing:0;font-family:var(--sans);}}
  .section-label .num{{display:inline-flex;align-items:center;justify-content:center;width:20px;height:20px;background:var(--unblue);color:#fff;border-radius:50%;font-size:10px;margin-right:8px;font-family:var(--sans);}}
  .head{{background:linear-gradient(120deg,var(--unblue-deep),var(--unblue) 75%);color:#f3ede0;padding:26px 34px 20px;position:relative;}}
  .head::before{{content:"";position:absolute;right:-40px;top:-40px;width:220px;height:220px;border-radius:50%;border:26px solid rgba(243,237,224,.07);}}
  .eyebrow{{font-size:10.5px;text-transform:uppercase;letter-spacing:.16em;opacity:.75;}}
  .headline{{font-family:var(--serif);font-weight:700;font-size:25px;line-height:1.15;max-width:580px;margin:6px 0 10px;}}
  .idbar{{display:flex;gap:20px;flex-wrap:wrap;font-size:11px;}}
  .idbar b{{display:block;font-size:13px;}}
  .idbar span{{opacity:.7;font-size:9.5px;text-transform:uppercase;letter-spacing:.06em;}}
  .verdict{{display:flex;background:var(--unblue-deep);color:#fff;padding:10px 34px;align-items:center;gap:14px;font-size:12px;flex-wrap:wrap;}}
  .verdict .big{{font-family:var(--serif);font-size:19px;font-weight:700;color:var(--amber);}}
  .funded-line{{font-size:11.5px;color:var(--ink-soft);margin:-2px 0 12px;font-style:italic;}}
  .funded-line b{{color:var(--unblue-deep);font-style:normal;}}
  .metarow{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:12px;}}
  .meta{{background:var(--panel);border:1px solid var(--line);border-radius:5px;padding:10px 10px 9px;text-align:center;}}
  .meta .mk{{font-size:8.5px;text-transform:uppercase;letter-spacing:.05em;color:var(--ink-soft);margin-bottom:4px;}}
  .meta .mv{{font-family:var(--serif);font-size:14.5px;font-weight:700;color:var(--unblue-deep);line-height:1.2;}}
  .meta.warn .mv{{color:var(--brick);}}
  .tlwrap{{position:relative;height:76px;margin:6px 0 4px;}}
  .tlline{{position:absolute;left:0;right:0;top:36px;height:3px;background:var(--line);}}
  .tlpoint{{position:absolute;top:26px;width:12px;height:12px;border-radius:50%;border:2px solid var(--paper);}}
  .tlpoint.plan{{background:var(--ink-soft);}}
  .tlpoint.actual{{background:var(--amber);}}
  .tllabel{{position:absolute;top:46px;font-size:8.5px;width:92px;text-align:center;color:var(--ink-soft);transform:translateX(-50%);}}
  .tlyear{{position:absolute;top:6px;font-size:9px;font-weight:700;color:var(--unblue-deep);transform:translateX(-50%);}}
  .tl-legend{{display:flex;gap:16px;font-size:9px;color:var(--ink-soft);margin-bottom:6px;}}
  .tl-legend span{{display:inline-flex;align-items:center;gap:4px;}}
  .tl-legend i{{width:8px;height:8px;border-radius:50%;display:inline-block;}}
  .delaywrap{{background:var(--panel);border:1px solid var(--line);border-radius:6px;padding:14px 16px;}}
  .ganttrow{{display:flex;align-items:center;gap:10px;margin-bottom:10px;}}
  .gantt-label{{width:70px;font-size:10px;color:var(--ink-soft);}}
  .gantt-track{{flex:1;background:#efe9d8;border-radius:3px;height:16px;position:relative;overflow:visible;}}
  .gantt-fill{{position:absolute;top:0;height:100%;border-radius:3px;}}
  .gantt-fill.plan{{background:#8fa4b5;}}
  .gantt-fill.overrun{{background:repeating-linear-gradient(45deg,var(--brick),var(--brick) 4px,#c96a4d 4px,#c96a4d 8px);}}
  .overrun-badge{{display:inline-block;background:var(--brick);color:#fff;font-size:11px;font-weight:700;padding:4px 10px;border-radius:3px;margin-top:6px;}}
  .delay-causes{{margin-top:12px;font-size:10.5px;}}
  .delay-causes li{{margin-bottom:4px;}}
  .twocol{{display:grid;grid-template-columns:1fr 1fr;gap:16px;}}
  .chartcard{{background:var(--panel);border:1px solid var(--line);border-radius:6px;padding:12px;}}
  .chartcard b{{font-size:10.5px;color:var(--unblue-deep);display:block;margin-bottom:6px;}}
  .chartcard canvas{{max-height:175px;}}
  .finbar{{display:flex;align-items:center;gap:10px;margin-bottom:8px;}}
  .finlabel{{width:82px;font-size:10px;color:var(--ink-soft);}}
  .finbg{{flex:1;background:#efe9d8;border-radius:3px;height:14px;overflow:hidden;}}
  .finfill{{height:100%;background:linear-gradient(90deg,var(--ozone),#3f9c8d);}}
  .finval{{font-size:10px;font-weight:700;color:var(--unblue-deep);width:60px;text-align:right;}}
  .stktable{{width:100%;border-collapse:collapse;font-size:10px;}}
  .stktable th{{background:var(--unblue);color:#fff;padding:6px 8px;text-align:left;font-weight:600;font-size:9.5px;}}
  .stktable td{{padding:5.5px 8px;border-bottom:1px solid var(--line);vertical-align:middle;}}
  .stktable tr:nth-child(even) td{{background:#faf8f2;}}
  .gbar{{display:flex;align-items:center;gap:6px;}}
  .gbar-bg{{width:60px;background:#efe9d8;border-radius:3px;height:8px;overflow:hidden;}}
  .gbar-fill{{height:100%;background:var(--brick);}}
  .gbar-val{{font-size:9.5px;width:30px;}}
  footer{{margin:18px 34px 0;padding-top:10px;border-top:1px solid var(--line);font-size:8.5px;color:var(--ink-soft);display:flex;justify-content:space-between;}}
  @media print{{body{{background:#fff;}}.page{{box-shadow:none;margin:0;}}}}
</style>
</head>
<body>
<div class="page">
  <div class="head">
    <div class="eyebrow">UNIDO Independent Evaluation · Standardized Terminal Evaluation Infographic</div>
    <div class="headline">{t_esc}</div>
    <div class="idbar">{idbar_html}</div>
  </div>
  {verdict_html}
  <div class="body">
    <div class="section-label"><span><span class="num">1</span>At a Glance</span></div>
    {('<div class="funded-line">'+donor_line+'</div>') if donor_line.strip() else ""}
    <div class="metarow">
      <div class="meta"><div class="mk">Country</div><div class="mv">{_esc(d.get("country") or "—")}</div></div>
      <div class="meta"><div class="mk">Thematic Area</div><div class="mv">{_esc(d.get("thematic") or "—")}</div></div>
      <div class="meta"><div class="mk">Evaluation Type</div><div class="mv">{_esc(d.get("report_type") or "Terminal Evaluation")}</div></div>
      <div class="meta"><div class="mk">Year</div><div class="mv">{_esc(str(d.get("year") or "—"))}</div></div>
    </div>
    <div class="metarow" style="grid-template-columns:repeat(3,1fr);">
      <div class="meta"><div class="mk">Budget</div><div class="mv">{_esc(budget_disp)}</div></div>
      <div class="meta"><div class="mk">Actual Duration</div><div class="mv">{_esc(_fmt_m(d.get("actual_months") or d.get("planned_months")))}</div></div>
      <div class="{overrun_cls}"><div class="mk">Schedule Overrun</div><div class="mv">{_esc(overrun_val)}</div></div>
    </div>
    {"" if not has_timeline else '<div class="section-label"><span><span class="num">2</span>Implementation Timeline</span></div><div class="tl-legend"><span><i style="background:var(--ink-soft);"></i>Planned milestone</span><span><i style="background:var(--amber);"></i>Actual milestone</span></div><div class="tlwrap" id="timeline"></div>'}
    {"" if not has_gantt else f'<div class="section-label"><span><span class="num">3</span>Delays in the Project</span><span class="tag">planned vs. actual duration</span></div><div class="delaywrap"><div class="ganttrow"><div class="gantt-label">Planned</div><div class="gantt-track"><div class="gantt-fill plan" style="left:0;width:{p_pct}%;"></div></div></div><div class="ganttrow"><div class="gantt-label">Actual</div><div class="gantt-track"><div class="gantt-fill plan" style="left:0;width:{p_pct}%;"></div><div class="gantt-fill overrun" style="left:{p_pct}%;width:{ov_pct}%;"></div></div></div><span class="overrun-badge">+{d["delay_months"]} months overrun</span>{causes_html}</div>'}
    {"" if not has_financial else f'<div class="section-label"><span><span class="num">{n4}</span>Financial Overview</span></div><div class="twocol"><div class="chartcard"><b>{spend_chart_title}</b><canvas id="spendChart"></canvas></div><div class="chartcard"><b>{cofin_chart_title}</b><canvas id="cofinChart"></canvas></div></div><div style="margin-top:12px;">{finbar_html}</div>'}
    <div class="section-label"><span><span class="num">{n5}</span>Stakeholder &amp; Gender Mainstreaming</span><span class="tag">{stk_tag}</span></div>
    <table class="stktable">
      <tr><th>Stakeholder Group</th><th>Role</th><th>Gender Status</th></tr>
      {stake_rows}
    </table>
    <div style="font-size:9.5px;color:var(--ink-soft);margin-top:8px;">
      <strong style="color:var(--unblue-deep);">Gender Mainstreaming — {_esc(g_rating)}:</strong> {_esc(g_note[:350])}
    </div>
  </div>
  <footer>
    <span>Source: UNIDO IEU Terminal Evaluation, {rid_esc} ({_esc(str(d.get("year") or ""))})</span>
    <span>Standardized infographic template · v1</span>
  </footer>
</div>
<script>
Chart.defaults.font.family="Segoe UI,Arial,sans-serif";
Chart.defaults.font.size=9.5;
Chart.defaults.color="#5a6b7a";
(function(){{
  const events={tl_events_json};
  const tl=document.getElementById('timeline');
  if(!tl||!events.length)return;
  tl.innerHTML='<div class="tlline"></div>';
  events.forEach(e=>{{
    tl.innerHTML+='<div class="tlpoint '+e.type+'" style="left:'+e.pos+'%;"></div>'+
      '<div class="tlyear" style="left:'+e.pos+'%;">'+e.yr+'</div>'+
      '<div class="tllabel" style="left:'+e.pos+'%;">'+e.lab+'</div>';
  }});
}})();
{financial_js_block}
</script>
</body>
</html>"""


def _build_report_infographic(rid: str) -> bytes:
    """Parse + render infographic. No LLM calls. Returns UTF-8 HTML bytes."""
    data = _parse_report_data(rid)
    return _render_infographic_html(data).encode("utf-8")


# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 — OECD-DAC Analysis
# ─────────────────────────────────────────────────────────────────────────────

def _make_dac_excel(report_ids: list, merged_evidence: dict, pilot_meta: dict) -> bytes:
    """Build a multi-sheet Excel export for the DAC analysis tab."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="003DA5")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    WRAP        = Alignment(wrap_text=True, vertical="top")

    RATING_SCALE_LOCAL = {
        6: "Highly Satisfactory (HS)",   5: "Satisfactory (S)",
        4: "Moderately Satisfactory (MS)", 3: "Moderately Unsatisfactory (MU)",
        2: "Unsatisfactory (U)",          1: "Highly Unsatisfactory (HU)",
    }
    CRITERIA_LOCAL = ["relevance", "effectiveness", "efficiency", "impact", "sustainability"]
    LABELS_LOCAL   = ["Relevance", "Effectiveness", "Efficiency", "Impact", "Sustainability"]

    def _style(ws, widths):
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 26
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
            for cell in row:
                cell.alignment = WRAP
            ws.row_dimensions[row_idx].height = max(
                15, min(15 * max(1, max((len(str(c.value or "")) for c in row), default=0) // 60), 120)
            )

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:

        # ── Sheet 1: DAC Ratings Summary ─────────────────────────────────
        rating_rows = []
        for rid in report_ids:
            p = pilot_meta.get(rid, {})
            dac_r  = p.get("dac_ratings", {})
            dac_rl = p.get("dac_rating_labels", {})
            row = {
                "Report ID":        rid,
                "Title":            p.get("title", ""),
                "Year":             p.get("year", ""),
                "Country":          p.get("country", ""),
                "Region":           p.get("region", ""),
                "Overall Rating":   f"{p.get('evaluation_rating','')}/6 — {p.get('overall_rating_label','')}",
            }
            for crit, lbl in zip(CRITERIA_LOCAL, LABELS_LOCAL):
                sc = dac_r.get(crit, "")
                full = dac_rl.get(crit, RATING_SCALE_LOCAL.get(sc, "")) if sc else ""
                row[lbl] = f"{sc}/6 — {full}" if sc else "—"
            rating_rows.append(row)
        df_rat = pd.DataFrame(rating_rows)
        df_rat.to_excel(writer, sheet_name="DAC Ratings", index=False)
        _style(writer.sheets["DAC Ratings"], [14, 55, 6, 18, 18, 28, 24, 24, 24, 24, 28])

        # ── Sheet 2: Evidence Passages ────────────────────────────────────
        ev_rows = []
        for crit, lbl in zip(CRITERIA_LOCAL, LABELS_LOCAL):
            for p in merged_evidence.get(crit, []):
                ev_rows.append({
                    "DAC Criterion":  lbl,
                    "Report Title":   p.get("report_title", ""),
                    "Year":           p.get("year", ""),
                    "Country":        p.get("country", ""),
                    "Evidence Passage": p.get("text", ""),
                })
        df_ev = pd.DataFrame(ev_rows) if ev_rows else pd.DataFrame(
            columns=["DAC Criterion", "Report Title", "Year", "Country", "Evidence Passage"])
        df_ev.to_excel(writer, sheet_name="Evidence Passages", index=False)
        _style(writer.sheets["Evidence Passages"], [22, 50, 6, 18, 100])

        # ── Sheet 3: Per-Criterion Detail ─────────────────────────────────
        for crit, lbl in zip(CRITERIA_LOCAL, LABELS_LOCAL):
            detail_rows = []
            for rid in report_ids:
                p   = pilot_meta.get(rid, {})
                sc  = p.get("dac_ratings", {}).get(crit, "")
                rl  = p.get("dac_rating_labels", {}).get(crit, "")
                evs = [e for e in merged_evidence.get(crit, [])
                       if e.get("country") == p.get("country")]
                detail_rows.append({
                    "Report":         p.get("title", rid),
                    "Country":        p.get("country", ""),
                    "Year":           p.get("year", ""),
                    f"{lbl} Rating":  f"{sc}/6 — {rl}" if sc else "—",
                    "Key Evidence":   " | ".join(e["text"][:300] for e in evs[:3]) if evs else "See findings section",
                })
            df_crit = pd.DataFrame(detail_rows)
            sheet_name = lbl[:31]  # Excel sheet name limit
            df_crit.to_excel(writer, sheet_name=sheet_name, index=False)
            _style(writer.sheets[sheet_name], [55, 18, 6, 28, 100])

    return buf.getvalue()


def _extract_dac_evidence_local(rid: str) -> dict:
    """
    Extract DAC criterion passages from local extracted_sections JSON.
    Priority: dedicated criterion sections → keyword-matched paragraphs from findings/conclusions.
    Returns dict: {criterion: [{"text":..., "report_title":..., "year":..., "country":...}]}
    """
    import re as _re
    sec_data = _load_sections_local(rid)
    secs     = sec_data.get("sections", {}) if sec_data else {}
    meta     = sec_data.get("metadata", {}) if sec_data else {}
    ai_d     = _load_ai_extraction(rid)
    ctx      = ai_d.get("context", {}) if ai_d else {}
    pilot    = PILOT_METADATA.get(rid, {})

    title   = pilot.get("title") or meta.get("title") or ctx.get("title") or rid
    year    = pilot.get("year")  or meta.get("year")  or ctx.get("year", "")
    country = pilot.get("country") or meta.get("country") or ctx.get("country", "")

    # Dedicated criterion section keys from ingestion pipeline
    SECTION_KEYS = {
        "relevance":      ["relevance"],
        "effectiveness":  ["effectiveness"],
        "efficiency":     ["efficiency"],
        "impact":         ["impact"],
        "sustainability": ["sustainability"],
    }

    # Fallback keyword patterns per criterion (applied to findings/conclusions text)
    KEYWORDS = {
        "relevance":      [r"relevan", r"national priorit", r"country priorit", r"strategic alignment", r"needs of", r"mandate"],
        "effectiveness":  [r"effectiv", r"achievement of", r"objectives.*achieved", r"results.*achieved", r"output.*delivered", r"target.*met"],
        "efficiency":     [r"efficien", r"cost.effective", r"value for money", r"implementation period", r"delay", r"overrun", r"budget utilisation"],
        "impact":         [r"impact", r"GHG reduction", r"emission", r"transformative", r"systemic change", r"long.term change", r"broader.*change"],
        "sustainability": [r"sustainab", r"after.*project", r"beyond.*project", r"institutional.*risk", r"financial.*risk", r"replicat", r"ownership"],
    }

    evidence = {c: [] for c in DAC_CRITERIA}

    def _make_passage(text: str, source: str = "") -> dict:
        return {"text": text[:750].strip(), "report_title": title,
                "year": year, "country": country, "source": source}

    # Step 1: use dedicated extracted sections if available
    for crit, keys in SECTION_KEYS.items():
        for k in keys:
            raw = secs.get(k, "")
            if raw and len(raw.strip()) > 100:
                # Split into paragraphs
                paras = [p.strip() for p in _re.split(r"\n{2,}", raw) if len(p.strip()) > 80]
                for para in paras[:5]:
                    evidence[crit].append(_make_passage(para, "dedicated section"))

    # Step 2: keyword search across findings + conclusions for any criterion still empty
    combined = "\n\n".join(filter(None, [secs.get("findings", ""), secs.get("conclusions", "")]))
    paragraphs = [p.strip() for p in _re.split(r"\n{2,}", combined) if len(p.strip()) > 100]

    for para in paragraphs:
        para_lower = para.lower()
        for crit, patterns in KEYWORDS.items():
            if len(evidence[crit]) >= 6:
                continue  # already have enough
            matches = sum(1 for pat in patterns if _re.search(pat, para_lower))
            if matches >= 1:
                # avoid duplicating passages already added
                if not any(p["text"][:80] == para[:80] for p in evidence[crit]):
                    evidence[crit].append(_make_passage(para, "findings/conclusions"))

    # Cap at 6 per criterion
    for crit in evidence:
        evidence[crit] = evidence[crit][:6]

    return evidence


def show_dac_tab():
    load_reports()
    # Demo phase: restrict to the 4 verified pilot reports only
    all_reps = [r for r in (st.session_state.all_reports or [])
                if r.get("report_id") in PILOT_METADATA]

    if not all_reps:
        st.info("No reports indexed yet.")
        return

    st.markdown(
        "Cross-portfolio analysis across the 5 OECD-DAC evaluation criteria. "
        "Select reports to compare their evidence coverage and browse verbatim passages "
        "across Relevance, Effectiveness, Efficiency, Impact, and Sustainability."
    )

    col_left, col_right = st.columns([3, 7])

    # ── Report selector ──────────────────────────────────────────────────────
    with col_left:
        st.markdown("#### Select Reports")
        mode = st.radio("Mode", ["Single report", "Compare multiple"], key="dac_mode",
                        horizontal=True)

        titles = {r["report_id"]: f"{r.get('year','')} — {r.get('title','')[:50]}"
                  for r in all_reps}

        if mode == "Single report":
            sel_id = st.selectbox("Report", list(titles.keys()),
                                  format_func=lambda x: titles.get(x, x),
                                  key="dac_single")
            dac_report_ids = [sel_id] if sel_id else []
        else:
            selected = st.multiselect("Reports (up to 5)", list(titles.keys()),
                                      format_func=lambda x: titles.get(x, x),
                                      max_selections=5, key="dac_multi")
            dac_report_ids = selected

        analyse = st.button("Analyse", type="primary", use_container_width=True)

    # ── Analysis panel ────────────────────────────────────────────────────────
    with col_right:
        if not dac_report_ids:
            st.info("Select a report on the left and click Analyse.")
            return

        # Trigger or use cached
        if analyse:
            with st.spinner("Extracting DAC evidence from reports…"):
                evidence_by_rep = {}
                for rid in dac_report_ids:
                    evidence_by_rep[rid] = _extract_dac_evidence_local(rid)
                st.session_state["dac_evidence_by_rep"] = evidence_by_rep
                st.session_state["dac_report_ids"]      = dac_report_ids
                st.session_state.pop("dac_ai_summary", None)

        if not st.session_state.get("dac_evidence_by_rep"):
            st.info("Click **Analyse** to extract DAC evidence from the selected reports.")
            return

        evidence_by_rep = st.session_state["dac_evidence_by_rep"]
        cached_ids      = st.session_state.get("dac_report_ids", dac_report_ids)

        # Merge evidence across all selected reports for the browser
        merged_evidence = {c: [] for c in DAC_CRITERIA}
        for rid in cached_ids:
            ev = evidence_by_rep.get(rid, {})
            for c in DAC_CRITERIA:
                merged_evidence[c].extend(ev.get(c, []))

        # ── Rating scale reference ─────────────────────────────────────────
        RATING_SCALE = {6: "HS", 5: "S", 4: "MS", 3: "MU", 2: "U", 1: "HU"}
        RATING_LABELS_FULL = {
            6: "Highly Satisfactory", 5: "Satisfactory",
            4: "Moderately Satisfactory", 3: "Moderately Unsatisfactory",
            2: "Unsatisfactory", 1: "Highly Unsatisfactory",
        }
        RATING_COLORS = {6: "#15803d", 5: "#16a34a", 4: "#ca8a04", 3: "#ea580c", 2: "#dc2626", 1: "#991b1b"}

        # ── Radar chart using actual DAC ratings ──────────────────────────
        st.markdown("#### OECD-DAC Criterion Ratings")
        st.caption("Ratings on the 6-point GEF/UNIDO scale: HS=6 · S=5 · MS=4 · MU=3 · U=2 · HU=1")

        TRACE_COLORS = ["#009EDB", "#f97316", "#22c55e", "#a855f7", "#ef4444"]
        fig_radar = go.Figure()

        for i, rid in enumerate(cached_ids):
            pilot = PILOT_METADATA.get(rid, {})
            dac_r = pilot.get("dac_ratings", {c: 3 for c in DAC_CRITERIA})
            scores = [dac_r.get(c, 3) for c in DAC_CRITERIA]
            rep_title = display_title(pilot) if pilot else rid
            country   = pilot.get("country", "")
            color     = TRACE_COLORS[i % len(TRACE_COLORS)]

            fig_radar.add_trace(go.Scatterpolar(
                r=scores + [scores[0]],
                theta=DAC_LABELS + [DAC_LABELS[0]],
                name=f"{country} ({pilot.get('year','')})",
                fill="toself",
                line=dict(color=color, width=2.5),
                opacity=0.7,
                hovertemplate=(
                    "<b>%{theta}</b><br>"
                    "Rating: %{r}/6<br>"
                    "<extra>" + country + "</extra>"
                ),
            ))

        fig_radar.update_layout(
            polar=dict(
                bgcolor="rgba(248,250,252,0.8)",
                radialaxis=dict(
                    visible=True,
                    range=[0, 6],
                    tickvals=[1, 2, 3, 4, 5, 6],
                    ticktext=["HU·1", "U·2", "MU·3", "MS·4", "S·5", "HS·6"],
                    tickfont=dict(size=9, color="#64748b"),
                    gridcolor="#e2e8f0",
                    linecolor="#cbd5e1",
                ),
                angularaxis=dict(
                    tickfont=dict(size=12, color="#1e3a5f", family="Inter, sans-serif"),
                    linecolor="#cbd5e1",
                ),
            ),
            showlegend=len(cached_ids) > 1,
            legend=dict(font=dict(size=11), bgcolor="rgba(255,255,255,0.9)", bordercolor="#e2e8f0", borderwidth=1),
            margin=dict(t=20, b=20, l=40, r=40),
            height=420,
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
        )
        st.plotly_chart(fig_radar, use_container_width=True)

        # ── Per-criterion rating cards ─────────────────────────────────────
        if len(cached_ids) == 1:
            rid    = cached_ids[0]
            pilot  = PILOT_METADATA.get(rid, {})
            dac_r  = pilot.get("dac_ratings", {})
            dac_rl = pilot.get("dac_rating_labels", {})
            stat_cols = st.columns(5)
            for col, crit, label, color in zip(stat_cols, DAC_CRITERIA, DAC_LABELS, DAC_COLORS):
                score = dac_r.get(crit, "—")
                rlabel = dac_rl.get(crit, "")
                rc = RATING_COLORS.get(score, "#64748b") if isinstance(score, int) else "#64748b"
                col.markdown(
                    f'<div style="background:white;border:2px solid {rc};border-radius:10px;'
                    f'padding:12px 8px;text-align:center;box-shadow:0 1px 4px rgba(0,0,0,0.06);">'
                    f'<div style="font-size:26px;font-weight:800;color:{rc};">{score}</div>'
                    f'<div style="font-size:9px;font-weight:700;color:{rc};letter-spacing:0.5px;">'
                    f'{RATING_SCALE.get(score,"") if isinstance(score,int) else ""}</div>'
                    f'<div style="font-size:10px;color:#1e3a5f;font-weight:600;margin-top:4px;">{label}</div>'
                    f'<div style="font-size:8.5px;color:#64748b;margin-top:2px;line-height:1.3;">{rlabel}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
        else:
            # Multi-report: show comparison table
            st.markdown("**Ratings comparison**")
            header = "| Criterion | " + " | ".join(
                PILOT_METADATA.get(r, {}).get("country", r[:8]) for r in cached_ids
            ) + " |"
            sep = "|---|" + "---|" * len(cached_ids)
            rows = []
            for crit, label in zip(DAC_CRITERIA, DAC_LABELS):
                cells = []
                for r in cached_ids:
                    sc = PILOT_METADATA.get(r, {}).get("dac_ratings", {}).get(crit, "—")
                    lbl = RATING_SCALE.get(sc, "—") if isinstance(sc, int) else "—"
                    cells.append(f"**{sc}** ({lbl})")
                rows.append(f"| {label} | " + " | ".join(cells) + " |")
            st.markdown("\n".join([header, sep] + rows))

        st.divider()

        # ── AI Summary ────────────────────────────────────────────────────
        st.markdown("#### AI Summary by Criterion")
        if st.button("Generate AI Summary across criteria", type="primary",
                     use_container_width=True, key="dac_ai_btn"):
            with st.spinner("Claude is analysing DAC evidence across selected reports…"):
                try:
                    import anthropic as _anthropic
                    try:
                        _api_key = st.secrets["ANTHROPIC_API_KEY"]
                    except Exception:
                        _api_key = os.getenv("ANTHROPIC_API_KEY", "")

                    if not _api_key:
                        st.warning("Add ANTHROPIC_API_KEY to Streamlit secrets.", icon="⚠️")
                    else:
                        context_lines = []
                        for crit, label in zip(DAC_CRITERIA, DAC_LABELS):
                            passages = merged_evidence.get(crit, [])[:4]
                            if passages:
                                context_lines.append(f"\n### {label.upper()}")
                                for p in passages:
                                    context_lines.append(
                                        f"[{p.get('report_title','')} – {p.get('country','')}]: "
                                        f"{p.get('text','')[:500]}"
                                    )
                        context_text = "\n".join(context_lines)

                        _client = _anthropic.Anthropic(api_key=_api_key)
                        msg = _client.messages.create(
                            model="claude-sonnet-4-5",
                            max_tokens=2048,
                            system=(
                                "You are a senior UNIDO evaluation analyst with deep expertise in "
                                "OECD-DAC evaluation criteria. You write concise, evidence-based "
                                "analytical summaries for UN senior management.\n\n"
                                "RULES:\n"
                                "1. Address each of the 5 DAC criteria with a clear heading\n"
                                "2. Cite specific reports by name\n"
                                "3. Be analytical — identify patterns and cross-cutting findings\n"
                                "4. Never invent information not in the provided evidence\n"
                                "5. End with ## Key Cross-Cutting Findings (3 bullet points)"
                            ),
                            messages=[{
                                "role": "user",
                                "content": (
                                    f"Analyse the OECD-DAC criteria evidence from these "
                                    f"{len(cached_ids)} UNIDO evaluation report(s) and provide "
                                    f"an analytical summary per criterion.\n\nEVIDENCE:\n{context_text}"
                                )
                            }]
                        )
                        st.session_state["dac_ai_summary"] = msg.content[0].text if msg.content else ""
                except Exception as e:
                    st.error(f"AI summary failed: {e}")

        if st.session_state.get("dac_ai_summary"):
            st.markdown(
                '<div style="background:white;border:1px solid #e5e7eb;border-radius:8px;'
                'padding:1rem 1.2rem;margin:0.5rem 0 1rem;font-size:0.87rem;line-height:1.7;">',
                unsafe_allow_html=True,
            )
            st.markdown(st.session_state["dac_ai_summary"])
            st.markdown('</div>', unsafe_allow_html=True)
            if st.button("Clear summary", key="dac_clear_sum"):
                del st.session_state["dac_ai_summary"]
                st.rerun()

        st.divider()

        # ── Evidence browser ───────────────────────────────────────────────
        browser_hdr, dl_col = st.columns([6, 2])
        with browser_hdr:
            st.markdown("#### Evidence Browser")
            st.caption("Verbatim passages extracted from report sections, matched to each DAC criterion.")
        with dl_col:
            try:
                xl_dac = _make_dac_excel(cached_ids, merged_evidence, PILOT_METADATA)
                label_str = "_".join(
                    PILOT_METADATA.get(r, {}).get("country", r[:6]) for r in cached_ids
                )
                st.download_button(
                    "⬇ Download Excel",
                    data=xl_dac,
                    file_name=f"UNIDO_DAC_Analysis_{label_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary",
                )
            except Exception as _xe:
                st.caption(f"Export error: {_xe}")

        tab_labels = [
            f"{lbl} ({len(merged_evidence.get(c, []))})"
            for c, lbl in zip(DAC_CRITERIA, DAC_LABELS)
        ]
        tabs = st.tabs(tab_labels)

        for tab, criterion, label, color in zip(tabs, DAC_CRITERIA, DAC_LABELS, DAC_COLORS):
            with tab:
                passages = merged_evidence.get(criterion, [])
                if not passages:
                    st.caption("No passages found for this criterion in the selected reports.")
                    continue
                for p in passages[:12]:
                    st.markdown(
                        f'<div class="dac-card">'
                        f'<div class="passage-title" style="color:{color};">'
                        f'{p.get("report_title","")}</div>'
                        f'<div class="passage-meta">'
                        f'{p.get("year","")} · {p.get("country","")}</div>'
                        f'<div class="dac-quote">{p.get("text","")[:650]}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

# ─────────────────────────────────────────────────────────────────────────────
# Admin tab
# ─────────────────────────────────────────────────────────────────────────────

def show_admin_tab():
    st.markdown("## User Management")
    st.markdown(
        "You are the administrator. Add, deactivate, or remove users. "
        "Changes take effect immediately."
    )

    try:
        resp = api("GET", "/api/v1/admin/users")
        data = resp.json() if resp.status_code == 200 else {}
    except Exception as e:
        st.error(str(e)); return

    users   = data.get("users", [])
    sessions= data.get("active_sessions", [])

    st.markdown(f"### Users ({len(users)}/6 slots)")
    for u in users:
        c1, c2, c3, c4 = st.columns([4, 1, 1, 1])
        with c1:
            rc = "role-admin" if u["role"] == "admin" else "role-user"
            ac = "color:#22c55e" if u["active"] else "color:#ef4444"
            st.markdown(
                f'**{u["display_name"]}** (@{u["username"]})<br>'
                f'<span class="role-badge {rc}">{u["role"]}</span> '
                f'<span style="font-size:.78rem;{ac}">{"● Active" if u["active"] else "● Inactive"}</span>',
                unsafe_allow_html=True,
            )
        with c2:
            st.caption(u.get("created_at","")[:10])
        with c3:
            if u["username"] != st.session_state.username:
                lbl = "Deactivate" if u["active"] else "Activate"
                if st.button(lbl, key=f"tgl_{u['username']}"):
                    r = api("PATCH", f"/api/v1/admin/users/{u['username']}/active",
                            json={"active": not u["active"]})
                    if r.status_code == 200:
                        st.success(r.json()["detail"]); st.rerun()
                    else:
                        st.error(r.json().get("detail","Error"))
        with c4:
            if u["username"] != st.session_state.username:
                if st.button("", key=f"del_{u['username']}"):
                    r = api("DELETE", f"/api/v1/admin/users/{u['username']}")
                    if r.status_code == 200:
                        st.success(r.json()["detail"]); st.rerun()
                    else:
                        st.error(r.json().get("detail","Error"))
        st.divider()

    if sessions:
        st.markdown(f"### Active sessions ({len(sessions)})")
        for s in sessions:
            st.caption(
                f"• {{s['display_name']}} (@{{s['username']}}) — "
                f"logged in {s['login_time'][:19].replace('T',' ')} UTC"
            )

    st.divider()

    st.markdown("### Add new user")
    with st.form("add_user"):
        ca, cb = st.columns(2)
        with ca:
            nu = st.text_input("Username (lowercase, no spaces)", placeholder="sarah.jones")
            np = st.text_input("Temporary password (min 8 chars)", type="password")
        with cb:
            nd = st.text_input("Full name", placeholder="Sarah Jones")
            nr = st.selectbox("Role", ["user", "admin"])
        if st.form_submit_button("Create user", type="primary"):
            if not all([nu, np, nd]):
                st.error("All fields required.")
            elif len(np) < 8:
                st.error("Password must be ≥ 8 characters.")
            else:
                r = api("POST", "/api/v1/admin/users",
                        json={"username": nu, "password": np, "display_name": nd, "role": nr})
                if r.status_code == 200:
                    st.success(f" User '{nu}' created."); st.rerun()
                else:
                    st.error(r.json().get("detail","Error"))

    st.divider()
    st.markdown("### Reset a user's password")
    usernames = [u["username"] for u in users if u["username"] != st.session_state.username]
    if usernames:
        with st.form("reset_pw"):
            cx, cy = st.columns(2)
            with cx:
                pw_user = st.selectbox("User", usernames)
            with cy:
                new_pw = st.text_input("New password", type="password")
            if st.form_submit_button("Reset password"):
                if not new_pw or len(new_pw) < 8:
                    st.error("Password must be ≥ 8 characters.")
                else:
                    r = api("POST", f"/api/v1/admin/users/{pw_user}/password",
                            json={"new_password": new_pw})
                    if r.status_code == 200:
                        st.success(r.json()["detail"])
                    else:
                        st.error(r.json().get("detail","Error"))
    else:
        st.caption("No other users to reset.")

# ─────────────────────────────────────────────────────────────────────────────
# Main app shell
# ─────────────────────────────────────────────────────────────────────────────

def show_main_app():
    is_admin = st.session_state.role == "admin"

    st.markdown(f"""
    <div class="unido-header">
      <div>
        <span class="unido-logo">UNIDO</span>
        <span class="unido-sub">  |  Evaluation Intelligence Platform  ·  IEU / EIO</span>
      </div>
      <div class="user-chip"> {st.session_state.display_name}</div>
    </div>
    """, unsafe_allow_html=True)

    # ── Left: filters | Right: main content ─────────────────────────────────────────────
    _lcol, _rcol = st.columns([1, 4], gap="medium")

    with _lcol:
        st.markdown("#### Filters")
        with st.form(key="filter_form"):
            thematic_sel = st.multiselect(
                "Thematic Area", THEMATIC_AREAS, key="f_thematic",
            )
            region_sel = st.multiselect(
                "Region",
                ["Africa", "Asia", "Europe", "Latin America", "Middle East", "Global"],
                key="f_region",
            )
            eval_type_sel = st.multiselect(
                "Evaluation Type",
                ["Project Evaluation", "Strategic Evaluation", "Country Evaluation",
                 "Synthesis", "Reference Document"],
                key="f_eval_type",
            )
            yr_sel = st.selectbox(
                "Year", ["All years", 2025, 2024, 2023, 2022, 2021], key="f_year",
            )
            st.form_submit_button("Apply Filters", use_container_width=True, type="primary")

        filters = {
            "thematic":   thematic_sel,
            "sdgs":       [],
            "eval_type":  eval_type_sel,
            "region":     region_sel,
            "years":      [yr_sel] if yr_sel != "All years" else [],
            "year_min":   yr_sel if yr_sel != "All years" else None,
            "year_max":   yr_sel if yr_sel != "All years" else None,
            "dac":        [],
        }
        st.divider()
        if st.button("Sign out", use_container_width=True):
            do_logout()

    with _rcol:
        # ── Tabs ──────────────────────────────────────────────────────────────────────────────
        tab_names = ["Search & Browse", "Synthesis", "Visualize", "OECD-DAC"]
        if is_admin:
            tab_names.append("Admin")

        tabs = st.tabs(tab_names)

        with tabs[0]:
            show_search_tab(filters)
        with tabs[1]:
            show_synthesis_tab(filters)
        with tabs[2]:
            show_visualize_tab()
        with tabs[3]:
            show_dac_tab()
        if is_admin:
            with tabs[4]:
                show_admin_tab()

        st.divider()
        st.markdown(
            "<p style='text-align:center;color:#9ca3af;font-size:.72rem;'>"
            "UNIDO IEU Evaluation Intelligence Platform · Internal use only · "
            "Retrieved passages should be verified against source documents before formal citation."
            "</p>",
            unsafe_allow_html=True,
        )

    # ─────────────────────────────────────────────────────────────────────────────
    # Router
    # ─────────────────────────────────────────────────────────────────────────────

if not st.session_state.session_token:
    show_login_page()
else:
    show_main_app()
